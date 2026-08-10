#!/usr/bin/env python3
"""Measure HPC docling ("SOM program") OCR accuracy the standard way — Character
Error Rate — benchmarked against a local AI OCR (Apple Vision), into an Excel workbook.

WHY THIS SCRIPT EXISTS / WHAT IT MEASURES
-----------------------------------------
OCR accuracy is conventionally reported as **Character Error Rate (CER)** and
**Word Error Rate (WER)**: the edit (Levenshtein) distance between the engine's
output and a human ground-truth transcription, divided by the length of the
ground truth. Character accuracy = 1 - CER. This is the primary metric here.

  CER = levenshtein(ocr_chars, truth_chars) / len(truth_chars)
  WER = levenshtein(ocr_words, truth_words) / len(truth_words)
  character_accuracy = max(0, 1 - CER)

A second, complementary metric — **fact recall** — is also reported: of a set of
verifiable facts (dollar amounts, dates, statute cites, step fractions) on each
page, what fraction the engine captured. CER measures transcription fidelity
overall; fact recall measures whether the *specific data the coding pipeline
needs* survived. A page can have low CER but still drop one critical number, or
high CER from cosmetic noise that never touches a figure — so both are shown.

NORMALIZATION POLICY (applied identically to every engine + the ground truth
before scoring, and documented so the numbers are reproducible):
  • strip markdown structural syntax docling emits (# | * `) — layout markup, not
    a transcription error;
  • collapse all whitespace (incl. newlines) to single spaces;
  • keep original case and punctuation (those ARE real OCR errors).

GROUND TRUTH: 6 pages transcribed by eye from the source page images (below).
PER-PAGE OCR (each engine's ACTUAL deliverable, scored per page):
  • docling  — the whole-document deliverable (ocr_test_output/<id>.txt). Each
    page's rendering is extracted from it by best partial alignment to the
    ground-truth page. (A context-free single-page re-OCR is NOT used: docling
    transcribes isolated pages noticeably worse -- it drops header codes and
    reorders more -- so scoring that would understate the real deliverable.)
  • Apple Vision — its per-page output cache (cache/apple_vision_cache/<id>/pNNNN.txt).

CAVEAT: 6-page stratified spot-check, not an exhaustive audit; and CER of the OCR
text is distinct from accuracy of the *coded dataset* (that needs Phase 4).

Run after the per-page docling HPC job finishes.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from rapidfuzz.distance import Levenshtein
from rapidfuzz.fuzz import partial_ratio_alignment

WORK = Path(__file__).resolve().parents[1]
ACC = WORK / "cache" / "ocr_accuracy_pages"
DOCLING_WHOLE = WORK / "ocr_test_output"                # whole-doc docling deliverable
AV_CACHE = WORK / "cache" / "apple_vision_cache"        # per-page Apple Vision
AV_WHOLE = WORK / "cache" / "apple_vision_ocr"          # whole-doc Apple Vision
PAGE_MAP = ACC / "page_map.json"
OUT_XLSX = WORK / "ocr_test_output" / "ocr_accuracy_benchmark.xlsx"

# Verifiable facts per page (secondary "fact recall" metric).
FACTS = {
    "WA p12": ["$27.50", "$40,742", "$217.87", "$27.23", "$3,000", "$25 per hour", "187 days"],
    "WA p18": ["one (1) year", "thirty-six (36) months", "five (5) days", "10%", "1.5 times", "FMLA"],
    "SA p30": ["Class II, Step 5", "December 1, 2015", "1/6", "1/7", "1/8", "7.6.8", "(2) days"],
    "SA p110": ["$200", "44506", "Assembly Bill No. 1x", "1999", "20.1.7", "Stull Act"],
    "BIRM p30": ["1100", "Alabama Association of School Boards", "08/06/74", "04/21/81", "01/27/04", "02/23/10"],
    "BIRM p120": ["4001", "certificates of deposit", "SAFE", "41-14A-2", "Code of Alabama 1975", "4/22/08"],
}

# Ground-truth transcriptions read by eye from the page images.
GT = {
"WA p12": """For buildings with athletic trainers who have teams that progress past week 10 of the state football playoffs
Athletic Trainers that are splitting extended days can be paid for a total of 10 hours at a rate of $27.50
This will be available for each week a team moves in the quarterfinals, semi-finals, and state championship
Any Athletic Trainer that is not a certified staff member will follow the approved certified salary schedule for teachers to determine hourly rate of pay.
Year 1: $40,742/187 days = $217.87 per day/8 hours = $27.23
Non-Certified Athletic Trainers rate of pay will increase annually following the district salary schedule.
Each of the six large high schools will also have available $3,000 for additional support paid at the rate of $25 per hour. These employees may work at multiple schools
No changes or additions to supplemental salaries may be paid without being approved by the Association and the Board.
All persons assigned to the positions listed above shall be required to perform the duties associated with the positions listed and shall be paid according to this schedule. Assignment of positions requires prior Board approval and issuance of supplemental contracts. Activity and Advisor salary amounts will be reflected on a Supplemental Extra Duty Contract. Extended Contract salary amounts will be reflected on a Supplemental Extra Day Contract.
If increments or percentages are to be split or shared with two or more people, the following condition must first be met:
Individuals receiving less than a full increment as listed for the position in the negotiated agreement must agree in writing.
Page 12 Negotiated Agreement 2022-2023""",
"WA p18": """D. Upon retirement, all remaining personal leave days will be compensated at the rate of 1.5 times the established rate for short-term substitutes.
E. Personal leave may not be taken the last two weeks of school without permission of the Building Administrator.
F. No more than 10% of the Employees in any building may be gone on personal leave at one time.
ARTICLE VI: PARENTAL LEAVE
A. An Employee may request up to one (1) year leave of absence, without pay, for the purpose of raising a child during the first year after birth or first year after adoption placement. This leave shall be referred to as "Parental Leave of Absence".
B. Requests for parental leave of absence shall be submitted in writing, indicating the approximate beginning and ending date of the parental leave of absence.
C. Parental leave of absence may be given at any time, but must start consistent with one of the following:
i. School year break
ii. Semester break
D. Parental leave of absence must end, and the Employee return to work at the beginning of one of the following:
i. School year break
ii. Semester break
E. The District may, at its discretion, choose to waive the provisions of section C and D of this Article in an emergency situation.
F. With the exception of a Category 1 contracted Employee, a qualifying Employee may return to the District in a position for which they are qualified but is not guaranteed the same position nor assignment they left.
G. Employees qualify for parental leave once every thirty-six (36) months. If the Employee qualifies for Family Medical Leave Act (FMLA), the FMLA runs concurrently with parental leave.
ARTICLE VII: BEREAVEMENT LEAVE
A. The Board agrees to grant to each Employee up to five (5) days bereavement leave for each death in the immediate family (grandfather, grandmother, father, mother, brother, sister, husband, wife, resident of the Employee's immediate household, child, grandchild, niece, nephew, foster child, stepchild, expectant child, aunt, uncle, cousin, or same individuals related through marriage) for the purpose of attending services, travel, and/or emotional support of family.
Page 18 Negotiated Agreement 2022-2023""",
"SA p30": """7.5.6 Extended Work Year - Unit members whose regular assignment is extended beyond their normal work year shall be paid for the additional days at their regular daily rate and shall be credited with additional sick leave as indicated in 8.1.5.
7.6 Extra Service Assignments
7.6.1 Unless otherwise set forth, Extra Service Assignments shall be remunerated based on the product of the assigned factor times Class II, Step 5 of the regular salary schedule. All remuneration which is based on a percentage of a salary schedule step will be rounded off to the nearest dollar, except for the Hourly Rate Factors of 7.6.8. All Extra-Service Assignments included in Article VII, sec. 6 shall be increased effective on December 1, 2015, in accordance to Article VII language.
A. Any unit members who volunteer to teach beyond the number of class periods required in a "regular" teacher's schedule at their assigned school site shall be compensated as follows:
1. Unit members who are assigned to secondary positions shall receive a percentage equal to that of the extra class period(s) taught (6th period = 1/6; 7th period = 1/7; 8th period = 1/8) of the unit member's contract salary and shall receive that rate of pay for all authorized paid absences.
2. Unit members who volunteer to teach an extra period shall receive two (2) days additional accumulative sick leave days per year or a pro-ration thereof.
7.6.2 Any and all remuneration is based upon service beyond the unit member's duty day.
A. All remuneration for extra-service assignments shall be paid twice a year if a full-time assignment or at the end of the activity if less than a full year except for department chairpersons, bilingual stipends, and athletic directors, who shall be paid monthly as a part of their contract salary.
7.6.3 Department Chairperson (basis of Class II, Step 5)
26""",
"SA p110": """1. Develop new teacher training and support programs pursuant to Education Code section 44506 (c)(5). (Assembly Bill No. 1x, First Extraordinary Session, 1999).
2. Develop professional development for all teachers pursuant to Education Code section 44506(c)(4). (Assembly Bill No. 1x, First Extraordinary Session, 1999).
B. The SAEA and the District shall also jointly develop a budget annually utilizing the AB IX funding not budgeted for the Peer Assistance and Review Program, to develop and implement new teacher training and support programs and staff development for all teachers.
C. Such programs and budget shall be recommended to the Governing Board for consideration and adoption pursuant to its annual budgeting process.
D. SAEA appointees shall receive $200 for each meeting attended. Meetings shall occur outside the normal duty day. The parties agree to evaluate this process and pay arrangement on an annual basis and make adjustments accordingly, to meet the needs of the PAR program and joint committee members.
20.1.7 Peer Assistance and Review Implementation
The Peer Assistance and Review Program, as a replacement for the Mentor Teacher Program and as found in the negotiated agreement between the District and the Association, is intended to provide peer assistance to tenured participating teachers who are referred for unsatisfactory performance or who have volunteered. The Consulting teacher's role shall not replace or in any way erode the responsibility of the principal or assistant principal as evaluators as described in the Stull Act. The statute does not permit the Consulting Teacher or the PAR Panel to evaluate the Participating Teacher. Under the statute the District maintains the responsibility to evaluate teachers. This remains the role of the principal/assistant principal as evaluators.
106""",
"BIRM p30": """BIRMINGHAM ORGANIZATION AND ADMINISTRATION 1100
Board Member Training, Development and Self Evaluation
General - Birmingham Board of Education members shall pursue ongoing training to develop and enhance their knowledge and effectiveness as Board members and to improve Board governance and operations. Training will include participation in:
a. Orientation for newly elected or appointed school board members;
b. Training or consulting workshops for the local Board as a whole;
c. State or national school board association events addressing Board governance or operation, or other Board member development opportunities relating to leadership development, Board governance, or Board operations.
Source of Training and Report - The requirements of this policy will be satisfied by participation in training provided by the Alabama Association of School Boards or other sources considered knowledgeable in school board governance and leadership and approved by the Board. Board members will provide a report to the Board about training experiences at the next available Board meeting.
Board Self Evaluation - The Board believes that evaluation of its effectiveness will improve the Board's performance, exemplify the kind of constructive assessments the Board encourages for all school personnel and programs, and promote a spirit of teamwork throughout the school system. The Board will assess its own performance annually according to goals established by the Board. This information will be used by the Board each year in establishing the goals and objectives it will strive to accomplish during the subsequent year.
BIRMINGHAM CITY SCHOOLS APPROVED: 08/06/74 REVISED: 04/21/81 REVISED: 01/27/04 REVISED: 02/23/10""",
"BIRM p120": """FINANCE 4001
Cash Management
Policy Birmingham City Board of Education will maintain an effective system of cash management that anticipates cash needs and plans adequately to satisfy those needs.
General Cash is required to pay for all goods and services purchased by the Board and to meet future obligations as they come due. The disbursement of cash is a regular occurrence, and a sufficient level of cash should be kept available to meet these needs. Therefore, only cash necessary to meet anticipated expenditures plus a reasonable reserve for emergencies should be kept available. Any excess cash should be invested in instruments backed by government securities as specified by law.
Procedures Maximizing Return on Idle Funds
Cash not required for operations will be invested. The cash balances in all Birmingham City Board of Education central office bank accounts will be monitored by the Chief School Financial Officer and the Superintendent. Any amount accumulated over the target balance in the checking accounts will be transferred into the appropriate investment vehicle, such as C.D.s, money markets, repurchase agreements and/or Treasury bills/notes. An activity report shall be sent by the banking institution every time funds are transferred into or out of the investment accounts. A monthly investment report shall also be received which will be used to post monthly interest earned and to reconcile the investment accounts to the amounts recorded in the general ledger. Interest rates will be monitored on the investment accounts and if other investment vehicles are earning higher interest rates, funds will be transferred from the investment account into the higher yield investment vehicle.
The Board shall purchase certificates of deposit (CDs) on a competitive basis. Competitive quotes for CDs maturing in one year or less on fixed rate CDs and two years or less on floating rate CDs will be obtained from qualified financial institutions. The Chief School Financial Officer and the Superintendent will choose what best fits the Board's portfolio based on the maturity required, liquidity requirements, the current makeup of the portfolio and the offered rate. The Chief School Financial Officer will provide the Board an investment report at least quarterly of the investment activity and interest earned.
In addition to the money market accounts, various government agency investment securities and other obligations of the U.S. government, such as Treasury Bills and Treasury Notes, shall be purchased by the investment official of the bank awarded the banking services of the Board. These shall be purchased with staggered maturity dates to maximize return and diversify the investment portfolio. Reports shall be sent by the bank upon purchases, interest payments and investment maturities. Upon receipt of these reports, the Chief School Financial Officer will ensure that journal entries are prepared to record the investment activity.
Collateralization:
All deposits of the Board must be secured by pledged collateral in an amount greater than the highest balance during any month, or deposited with a Qualified Public Depository of the Security for Alabama Funds Enhancement (SAFE) Program, according to Sections 41-14A-2 through 41-14A-6, 41-14A-8 and 41-14A-9, Code of Alabama 1975, as amended.
Birmingham City Schools Approved: 4/22/08""",
}


def normalize(text: str) -> str:
    """Normalization policy (see module docstring): drop markdown syntax, collapse
    whitespace, keep case + punctuation."""
    text = re.sub(r"[#|*`]", " ", text)
    text = re.sub(r"!\[[^\]]*\]\([^)]*\)", " ", text)   # any stray markdown image
    return re.sub(r"\s+", " ", text).strip()


def cer(ocr: str, truth: str) -> float:
    t = normalize(truth)
    return Levenshtein.distance(normalize(ocr), t) / max(len(t), 1)


def wer(ocr: str, truth: str) -> float:
    tw = normalize(truth).split()
    return Levenshtein.distance(normalize(ocr).split(), tw) / max(len(tw), 1)


def fact_norm(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[,;]", "", s.lower())).strip()


def _whole_norm(path: Path) -> str:
    if not path.exists():
        return ""
    raw = re.sub(r"!\[[^\]]*\]\(data:image/[^)]*\)", "", path.read_text(encoding="utf-8", errors="ignore"))
    return normalize(raw)


def docling_page_span(document_id: str, gt: str) -> str:
    """The docling deliverable is whole-document (no page breaks). Extract this
    page's rendering by best partial alignment of the ground-truth page against
    the whole-doc docling text, so CER reflects the ACTUAL deliverable -- not a
    context-free single-page re-OCR, which docling transcribes noticeably worse
    (it drops header codes and reorders more on isolated pages)."""
    whole = _whole_norm(DOCLING_WHOLE / f"{document_id}.txt")
    if not whole:
        return ""
    al = partial_ratio_alignment(normalize(gt), whole)
    return whole[al.dest_start:al.dest_end] if al else ""


def load_av_page(document_id: str, page: int) -> str:
    p = AV_CACHE / document_id / f"p{page:04d}.txt"
    return p.read_text(encoding="utf-8", errors="ignore") if p.exists() else ""


def main() -> None:
    pmap = json.loads(PAGE_MAP.read_text(encoding="utf-8"))
    rows = []            # per page: label, engine texts, metrics
    missing = []
    for page_uuid, info in pmap.items():
        label = info["label"]
        truth = GT.get(label, "")
        if not truth:
            print(f"WARNING: no ground truth for {label}"); continue
        dtext = docling_page_span(info["doc_id"], truth)
        atext = load_av_page(info["doc_id"], info["page"])
        if not dtext:
            missing.append(label)
        rows.append({
            "label": label, "truth": truth, "docling": dtext, "av": atext,
            "d_cer": cer(dtext, truth) if dtext else None,
            "d_wer": wer(dtext, truth) if dtext else None,
            "a_cer": cer(atext, truth) if atext else None,
            "a_wer": wer(atext, truth) if atext else None,
        })
    if missing:
        print(f"WARNING: per-page docling missing for {missing} — run the HPC job / wait for it.")

    def mean(vals):
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else None

    d_cer_m = mean(r["d_cer"] for r in rows); a_cer_m = mean(r["a_cer"] for r in rows)
    d_wer_m = mean(r["d_wer"] for r in rows); a_wer_m = mean(r["a_wer"] for r in rows)

    # Fact recall against each engine's WHOLE-DOC deliverable (does the fact survive anywhere).
    doc_of = {info["label"]: info["doc_id"] for info in pmap.values()}
    dwhole = {d: fact_norm(_whole_norm(DOCLING_WHOLE / f"{d}.txt")) for d in set(doc_of.values())}
    awhole = {d: fact_norm(_whole_norm(AV_WHOLE / f"{d}.txt")) for d in set(doc_of.values())}
    fr = {"docling": [0, 0], "av": [0, 0]}
    fact_rows = []
    for r in rows:
        dn, an = dwhole[doc_of[r["label"]]], awhole[doc_of[r["label"]]]
        for tok in FACTS.get(r["label"], []):
            dh = fact_norm(tok) in dn
            ah = fact_norm(tok) in an
            fact_rows.append((r["label"], tok, dh, ah))
            fr["docling"][1] += 1; fr["av"][1] += 1
            if dh: fr["docling"][0] += 1
            if ah: fr["av"][0] += 1

    GREEN = PatternFill("solid", fgColor="C6EFCE"); RED = PatternFill("solid", fgColor="FFC7CE")
    YEL = PatternFill("solid", fgColor="FFEB9C")
    HDR = PatternFill("solid", fgColor="1F4E78"); HF = Font(color="FFFFFF", bold=True); BOLD = Font(bold=True)
    wb = openpyxl.Workbook()

    # ---- Summary ----
    su = wb.active; su.title = "Summary"
    def row(*vals, bold=False, fill=None, font=None):
        su.append(list(vals))
        for c in su[su.max_row]:
            if bold: c.font = BOLD
            if font: c.font = font
            if fill: c.fill = fill
    def pct(x): return "n/a" if x is None else f"{round(100*(1-x),1)}%"
    row("OCR accuracy — Character Error Rate (the standard OCR metric)", bold=True)
    row("Character accuracy = 1 - CER, vs page-image ground truth. 6-page stratified sample across the 3 docs.")
    row("")
    row("Engine", "Character accuracy", "Word accuracy", "CER", "WER", bold=True, fill=HDR)
    for c in su[su.max_row]: c.font = HF
    row("docling  (SOM / HPC — OCR of record)", pct(d_cer_m), pct(d_wer_m),
        "n/a" if d_cer_m is None else f"{round(100*d_cer_m,1)}%", "n/a" if d_wer_m is None else f"{round(100*d_wer_m,1)}%")
    row("Apple Vision  (local AI benchmark)", pct(a_cer_m), pct(a_wer_m),
        "n/a" if a_cer_m is None else f"{round(100*a_cer_m,1)}%", "n/a" if a_wer_m is None else f"{round(100*a_wer_m,1)}%")
    row("")
    row("Per-page character accuracy (1 - CER)", bold=True)
    row("Page", "docling", "Apple Vision", "docling CER", "AV CER", bold=True, fill=HDR)
    for c in su[su.max_row]: c.font = HF
    for r in rows:
        su.append([r["label"], pct(r["d_cer"]), pct(r["a_cer"]),
                   "n/a" if r["d_cer"] is None else f"{round(100*r['d_cer'],1)}%",
                   "n/a" if r["a_cer"] is None else f"{round(100*r['a_cer'],1)}%"])
    row("")
    dfp = round(100*fr["docling"][0]/max(fr["docling"][1],1),1); afp = round(100*fr["av"][0]/max(fr["av"][1],1),1)
    row("Secondary metric — fact recall (dollar amounts, dates, cites, fractions)", bold=True)
    row("docling", f"{fr['docling'][0]}/{fr['docling'][1]} = {dfp}%")
    row("Apple Vision", f"{fr['av'][0]}/{fr['av'][1]} = {afp}%")
    row("")
    row("Read as:", bold=True)
    if d_cer_m is not None:
        row(f"  • docling transcribed ~{round(100*(1-d_cer_m),1)}% of characters correctly (CER {round(100*d_cer_m,1)}%);")
        row(f"    the local benchmark ~{round(100*(1-a_cer_m),1)}%. Character accuracy is the standard OCR metric.")
    row("  • Fact recall (below) is the task-relevant complement: does the OCR capture the specific numbers/dates the coding uses.")
    row("")
    row("Caveats (for your supervisor):", bold=True)
    row("  • 6-page stratified spot-check, not an exhaustive audit.")
    row("  • CER normalizes whitespace + markdown layout (documented in the script); case & punctuation errors ARE counted.")
    row("  • OCR-text accuracy is distinct from CODED-DATASET accuracy — the latter needs the coding pipeline re-run (Phase 4).")
    for col, w in zip("ABCDE", (42, 18, 16, 14, 12)):
        su.column_dimensions[col].width = w

    # ---- Fact-recall detail ----
    fd = wb.create_sheet("Fact recall detail")
    fd.append(["Page", "Verifiable fact", "docling", "Apple Vision"])
    for c in fd[1]: c.fill, c.font = HDR, HF
    for label, tok, dh, ah in fact_rows:
        fd.append([label, tok, "✓" if dh else "MISS", "✓" if ah else "MISS"])
        fd.cell(fd.max_row, 3).fill = GREEN if dh else RED
        fd.cell(fd.max_row, 4).fill = GREEN if ah else RED
    for col, w in zip("ABCD", (12, 34, 12, 14)): fd.column_dimensions[col].width = w
    fd.freeze_panes = "A2"

    # ---- Per-page char detail ----
    pp = wb.create_sheet("Per-page detail")
    pp.append(["Page", "GT chars", "docling CER", "docling char-acc", "docling WER",
               "AV CER", "AV char-acc", "AV WER"])
    for c in pp[1]: c.fill, c.font = HDR, HF
    for r in rows:
        pp.append([r["label"], len(normalize(r["truth"])),
                   None if r["d_cer"] is None else round(r["d_cer"],4), pct(r["d_cer"]),
                   None if r["d_wer"] is None else round(r["d_wer"],4),
                   None if r["a_cer"] is None else round(r["a_cer"],4), pct(r["a_cer"]),
                   None if r["a_wer"] is None else round(r["a_wer"],4)])
    for col, w in zip("ABCDEFGH", (12,10,12,14,12,10,12,10)): pp.column_dimensions[col].width = w

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"docling: char-acc {pct(d_cer_m)}, WER-acc {pct(d_wer_m)}  |  "
          f"Apple Vision: char-acc {pct(a_cer_m)}, WER-acc {pct(a_wer_m)}")
    print(f"fact recall: docling {dfp}%  |  Apple Vision {afp}%")
    print(f"wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
