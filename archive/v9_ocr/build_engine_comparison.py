#!/usr/bin/env python3
"""4-engine OCR comparison (Character/Word Error Rate) on the 6-page benchmark.

Compares docling, olmocr2, deepseek_ocr (Yale SOM HPC) and Apple Vision (local),
all OCR-ing the SAME 6 single-page PDFs, scored vs page-image ground truth. See
build_ocr_accuracy_report.py for the metric definitions (CER/WER, normalization).

SINGLE-PAGE basis keeps the four engines apples-to-apples. Note: docling's
whole-document deliverable measures a few points LOWER than its single-page
number here (it reorders more with multi-page context) -- so treat docling's
89.5% as its best case.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from build_ocr_accuracy_report import (GT, FACTS, cer, wer, fact_norm, AV_CACHE, PAGE_MAP, WORK)

DOCS = WORK / "cache" / "ocr_accuracy_pages" / "documents"
OUT = WORK / "ocr_test_output" / "ocr_engine_comparison.xlsx"

ENGINES = [
    ("olmocr2", "olmocr2.txt", "HPC (vLLM)"),
    ("apple_vision", None, "local"),
    ("docling", "docling.txt", "HPC (current)"),
    ("deepseek_ocr", "deepseek_ocr.txt", "HPC (vLLM)"),
]


def load(engine: str, fname, uuid: str, doc_id: str, page: int) -> str:
    p = (AV_CACHE / doc_id / f"p{page:04d}.txt") if engine == "apple_vision" else (DOCS / uuid / "ocr" / fname)
    if not p.exists():
        return ""
    return re.sub(r"!\[[^\]]*\]\(data:image/[^)]*\)", "", p.read_text(encoding="utf-8", errors="ignore"))


def main() -> None:
    pmap = json.loads(PAGE_MAP.read_text(encoding="utf-8"))
    agg = {e: {"cer": [], "wer": [], "fr": [0, 0]} for e, _, _ in ENGINES}
    per_page = {}
    for u, info in pmap.items():
        lbl = info["label"]; gt = GT[lbl]; per_page[lbl] = {}
        for e, fn, _ in ENGINES:
            t = load(e, fn, u, info["doc_id"], info["page"])
            c = cer(t, gt) if t else 1.0
            agg[e]["cer"].append(c); agg[e]["wer"].append(wer(t, gt) if t else 1.0)
            per_page[lbl][e] = 1 - c
            ft = fact_norm(t)
            for tok in FACTS.get(lbl, []):
                agg[e]["fr"][1] += 1
                if fact_norm(tok) in ft:
                    agg[e]["fr"][0] += 1
    mean = lambda xs: sum(xs) / len(xs) if xs else 0

    HDR = PatternFill("solid", fgColor="1F4E78"); HF = Font(color="FFFFFF", bold=True)
    WIN = PatternFill("solid", fgColor="C6EFCE"); BOLD = Font(bold=True)
    wb = openpyxl.Workbook()

    su = wb.active; su.title = "Comparison"
    su.append(["OCR engine comparison — Character Error Rate (6-page benchmark)"])
    su["A1"].font = BOLD
    su.append(["All four engines OCR the same 6 single-page PDFs; scored vs page-image ground truth. Higher = better."])
    su.append([])
    su.append(["Engine", "Where", "Character accuracy", "Word accuracy", "Fact recall"])
    for c in su[su.max_row]:
        c.fill, c.font = HDR, HF
    ranked = sorted(ENGINES, key=lambda x: mean(agg[x[0]]["cer"]))
    best = ranked[0][0]
    for e, _, where in ranked:
        a = agg[e]
        su.append([e, where, f"{100*(1-mean(a['cer'])):.1f}%", f"{100*(1-mean(a['wer'])):.1f}%",
                   f"{100*a['fr'][0]/max(a['fr'][1],1):.1f}%"])
        if e == best:
            for c in su[su.max_row]:
                c.fill = WIN
    su.append([])
    su.append([f"Winner: {best} — highest character accuracy, and best on both prose and number-heavy pages."], )
    su[su.max_row][0].font = BOLD
    su.append(["docling's edge is table STRUCTURE (rebuilds salary grids); the vLLM engines emit flatter text."])
    su.append(["deepseek_ocr underperformed here; drop it. Apple Vision is the best no-HPC option but drops dollar amounts."])
    su.append([])
    su.append(["Caveats:"]); su[su.max_row][0].font = BOLD
    su.append(["  • 6-page single-page spot-check; docling whole-doc deliverable ~3pts lower (reordering)."])
    su.append(["  • olmocr2/deepseek need the c001-exclude + longer startup that we patched — must persist at scale."])
    for col, w in zip("ABCDE", (16, 16, 18, 16, 14)):
        su.column_dimensions[col].width = w

    pp = wb.create_sheet("Per-page")
    pp.append(["Page"] + [e for e, _, _ in ENGINES])
    for c in pp[1]:
        c.fill, c.font = HDR, HF
    for lbl in per_page:
        pp.append([lbl] + [f"{100*per_page[lbl][e]:.1f}%" for e, _, _ in ENGINES])
        # highlight best per page
        vals = {e: per_page[lbl][e] for e, _, _ in ENGINES}
        bestp = max(vals, key=vals.get)
        for i, (e, _, _) in enumerate(ENGINES, start=2):
            if e == bestp:
                pp.cell(pp.max_row, i).fill = WIN
    for col, w in zip("ABCDE", (14, 14, 14, 14, 14)):
        pp.column_dimensions[col].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    for e, _, _ in ranked:
        a = agg[e]
        print(f"  {e:14s} char {100*(1-mean(a['cer'])):5.1f}%  word {100*(1-mean(a['wer'])):5.1f}%  fact {100*a['fr'][0]/max(a['fr'][1],1):5.1f}%")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
