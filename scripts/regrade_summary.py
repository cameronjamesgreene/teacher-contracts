#!/usr/bin/env python3
"""Rewrite 1_Summary for the three-level re-grade, in place.

Edits the existing sheet rather than rebuilding it, so hand edits to the summary survive
(the colour legend the user deleted stays deleted). Adds the % Strict column back — it
was removed when the audit was binary and Strict merely repeated Acceptable — and states
the v9 comparison the whole re-grade exists to support.
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import OUT_DIR
from regrade_v9_scale import FILL, estimate

PROGRAMS = [("2_llm_extract", "LLM extract (questions)"),
            ("3_salary_schedule", "Salary schedule (grids)"),
            ("4_rights_score", "Rights (clauses)")]
# v9's own figures, recomputed from the COUNTS in output/output_v9/audit_report_v9.xlsx
# rows 12-14 rather than its rounded percentages, so both versions are derived the same
# way: llm 68/7/4 of 79, salary 140/27/8 of 175 scored, rights 64/11/0 of 75.
V9 = {"LLM extract (questions)": (68 / 79, 75 / 79),
      "Salary schedule (grids)": (140 / 175, 167 / 175),
      "Rights (clauses)": (64 / 75, 75 / 75)}


def write_prose(sheet, row: int, text: str, *, bold=False, fill=None) -> None:
    cell = sheet.cell(row=row, column=1, value=text)
    cell.font = Font(bold=bold)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=OUT_DIR / "audit_report_v12.xlsx")
    args = parser.parse_args()
    workbook = load_workbook(args.workbook)
    summary = workbook["1_Summary"]

    collected, strata = {}, {}
    for name, label in PROGRAMS:
        sheet = workbook[name]
        header = [c.value for c in sheet[1]]
        rows = [{header[i]: c.value for i, c in enumerate(r) if i < len(header)}
                for r in sheet.iter_rows(min_row=2)]
        packed = [{"grade": r["grade_v9_scale"], "stratum": r.get("stratum", "?"),
                   "stratum_size": 1, "document": r.get("document_id") or r.get("pdf", "?")}
                  for r in rows]
        sizes: dict[str, int] = defaultdict(int)
        for r in rows:
            sizes[r.get("stratum", "?")] += 1
        collected[label] = (rows, packed)
        strata[label] = sizes

    # ── results table (header r6, data r7-r9) ────────────────────────────────────────
    headers = ["Program", "Sampled", "CORRECT", "PARTIAL", "INCORRECT", "% Acceptable",
               "% Strict", "Est. corpus (acceptable)", "95% CI low", "95% CI high",
               "Documents"]
    for column, text in enumerate(headers, start=1):
        cell = summary.cell(row=6, column=column, value=text)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL["header"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    from json import loads
    ledger = [loads(line) for line in
              (OUT_DIR / "audit_ledger.jsonl").read_text(encoding="utf-8").splitlines()
              if line.strip()]
    real_sizes = {(r["sheet"], r.get("stratum", "?")): r.get("stratum_size") or 1
                  for r in ledger}

    for index, (name, label) in enumerate(PROGRAMS):
        rows, packed = collected[label]
        for item in packed:
            item["stratum_size"] = real_sizes.get((name, item["stratum"]), 1)
        strict, acceptable = estimate(packed, False), estimate(packed, True)
        counts = defaultdict(int)
        for r in rows:
            counts[r["grade_v9_scale"]] += 1
        values = [label, len(rows), counts["CORRECT"], counts["PARTIAL"],
                  counts["INCORRECT"], acceptable["raw"], strict["raw"],
                  acceptable["weighted"], acceptable["low"], acceptable["high"],
                  strict["documents"]]
        for column, value in enumerate(values, start=1):
            cell = summary.cell(row=7 + index, column=column, value=value)
            cell.alignment = Alignment(vertical="top")
            if column in (6, 7):
                cell.number_format = "0%"
                cell.font = Font(bold=column == 6)
            elif column in (8, 9, 10):
                cell.number_format = "0.000"
            if column == 3:
                cell.fill = PatternFill("solid", fgColor=FILL["CORRECT"])
            if column == 4:
                cell.fill = PatternFill("solid", fgColor=FILL["PARTIAL"])
            if column == 5:
                cell.fill = PatternFill("solid", fgColor=FILL["INCORRECT"])

    write_prose(summary, 3,
        "Audit method: 1,470 rows checked. Every one of the 268 llm_extract absence "
        "claims was then RE-GRADED on v9's three-level scale by probing its document for "
        "the question bank's own published search terms (scripts/probe_absence_claims.py); "
        "125 returned no hit at all and 143 were read individually.")
    write_prose(summary, 11,
        "% Strict counts only rows judged fully correct. % Acceptable also counts PARTIAL. "
        "PARTIAL does NOT mean the same thing in each sheet, because v9 did not use one "
        "definition either: llm_extract = a not_discussed answer where the document holds "
        "material ADJACENT to the question; rights = the rule and the truth are "
        "right-vs-obligation, which Hohfeld treats as two sides of ONE relation, so the "
        "relation was found and only the perspective differs; salary = cell agreement in "
        "the 0.70-0.95 band. 'Est. corpus' is Horvitz-Thompson (each stratum's rate "
        "weighted by its true size); the interval is bootstrapped over DOCUMENTS, not rows.")

    write_prose(summary, 13, "COMPARISON WITH v9 — THE POINT OF THE RE-GRADE", bold=True)
    write_prose(summary, 14,
        "v9 scored 86% Strict / 95% Acceptable on llm_extract. Quoting v12's single rate "
        "against v9's STRICT figure overstated the gain; against its ACCEPTABLE figure it "
        "understated it. Below, both on one scale. NOTE the v12 columns are the "
        "CORPUS ESTIMATES, not the raw sample rates in the table above: v12 deliberately "
        "oversampled the error-prone strata (rule-vs-model disagreements are 33% of the "
        "rights sample against 14.7% of the corpus), so its raw rates are pulled down by "
        "design and are NOT comparable to v9's roughly representative sample. Comparing "
        "raw-to-raw here would repeat exactly the mistake this re-grade exists to fix.")
    for column, text in enumerate(["Program", "v9 Strict (sample)",
                                   "v12 Strict (corpus est.)", "v9 Accept. (sample)",
                                   "v12 Accept. (corpus est.)", "v9 corpus",
                                   "v12 corpus"], start=1):
        cell = summary.cell(row=15, column=column, value=text)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL["header"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for index, (name, label) in enumerate(PROGRAMS):
        rows, packed = collected[label]
        for item in packed:
            item["stratum_size"] = real_sizes.get((name, item["stratum"]), 1)
        strict, acceptable = estimate(packed, False), estimate(packed, True)
        v9s, v9a = V9[label]
        for column, value in enumerate([label, v9s, strict["weighted"], v9a,
                                        acceptable["weighted"],
                                        "15 docs, scans excluded",
                                        "42 docs, scans included"], start=1):
            cell = summary.cell(row=16 + index, column=column, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if column in (2, 3, 4, 5):
                cell.number_format = "0%"
                cell.font = Font(bold=column in (3, 5))
    write_prose(summary, 19,
        "READ THE llm_extract STRICT FIGURE AS AN UPPER BOUND. The re-grade's keyword "
        "probe is measurably less sensitive than reading: it missed 4 of the 9 absence "
        "errors the original audit had found by reading (Winston-Salem's page is mojibake, "
        "which defeats keyword search entirely). It also caught 2 hard errors the original "
        "missed - Milwaukee, whose contents page names an 'Outside Experience Credit' "
        "section, and Miami-Dade's 'Computer Network for Use By Teachers'. Two instruments "
        "finding 9 and 24 problems while overlapping on only 5 implies, by Chapman "
        "capture-recapture, about 41 genuinely non-correct absence rows against the 28 "
        "identified. Applying that correction puts llm_extract nearer 92% Strict / 97% "
        "Acceptable - still above v9 on both scales.")

    # ── per-stratum table (header r21, data r22-r32) ─────────────────────────────────
    for column, text in enumerate(["Program", "Stratum", "Audited",
                                   "Stratum size in corpus", "Strict in stratum",
                                   "Acceptable in stratum"], start=1):
        cell = summary.cell(row=21, column=column, value=text)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL["header"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    line = 22
    for name, label in PROGRAMS:
        rows, _ = collected[label]
        by: dict[str, list[str]] = defaultdict(list)
        for r in rows:
            by[r.get("stratum", "?")].append(r["grade_v9_scale"])
        for stratum in sorted(by):
            graded = [g for g in by[stratum] if g != "EXCLUDED"]
            if not graded:
                continue
            values = [label, stratum, len(graded),
                      real_sizes.get((name, stratum), len(graded)),
                      sum(1 for g in graded if g == "CORRECT") / len(graded),
                      sum(1 for g in graded if g in ("CORRECT", "PARTIAL")) / len(graded)]
            for column, value in enumerate(values, start=1):
                cell = summary.cell(row=line, column=column, value=value)
                cell.alignment = Alignment(vertical="top")
                if column in (5, 6):
                    cell.number_format = "0.000"
            line += 1

    write_prose(summary, 35,
        "- llm_extract remains the strongest program and improves on v9 on BOTH scales. "
        "Its residual errors are almost entirely WRONG ABSENCES: provisions the document "
        "names in a contents page or appendix that the extractor reported as not_discussed.")
    write_prose(summary, 37,
        "- rights_score: the deterministic Table A.3 rule is the weak component, and it is "
        "the one being shipped. Where rule and model disagree the model is right about "
        "three times in four, so 32 of the 89 INCORRECT rows are pure right-vs-obligation "
        "perspective (graded PARTIAL) and the rest are the rule losing outright. Resolving "
        "disagreements toward the model would move this program to roughly 0.95.")
    write_prose(summary, 38,
        "- salary_schedule: 46 grids place real values in the wrong cells - the numbers are "
        "printed on the page but land under the wrong step or lane. 51 of the 57 "
        "wrong-or-partial grids come from just five documents (Cleveland, LA, Pittsburgh, "
        "Fresno, 33.pdf).")
    write_prose(summary, 46,
        "- The salary PARTIAL band is NOT v9's. v9 marked a grid PARTIAL when its values "
        "were right but the grid was TRUNCATED (steps 1-38 captured, truth ran to 50). v12 "
        "cannot detect that at all: expected_cells is unpopulated for all 431 rows, capture "
        "measures share-of-page rather than share-of-table, and the cell-agreement check "
        "matches rows forward and stops - so a grid that ends early still scores perfectly. "
        "v12's salary figures are therefore optimistic against v9's by an unmeasured margin.")
    write_prose(summary, 47,
        "- Salary lane labels and covered employee groups were never verified - only values. "
        "A provision on a page nothing cited is invisible to this method.")

    workbook.save(args.workbook)
    print(f"  1_Summary rewritten -> {args.workbook}")


if __name__ == "__main__":
    main()
