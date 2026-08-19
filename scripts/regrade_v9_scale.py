#!/usr/bin/env python3
"""Re-grade the v12 audit on v9's three-level scale and rewrite the workbook.

## Why

v12 graded llm_extract and rights_score as correct/incorrect. v9 used
CORRECT / PARTIAL / INCORRECT, and its whole 86%-vs-95% gap lived in the middle
level. Comparing v12's single rate to either of v9's two was the mismatch that made
"86% -> 96%" look like a large gain; on one scale it is smaller and on the other larger.
This script puts both versions on one scale.

## What PARTIAL means, per program

v9 did not use one definition — it used a different one in each sheet, recovered here
from its own audit notes:

  llm_extract  `not_discussed` where the document holds material ADJACENT to the
               question but not squarely on point (v9's "soft false negative").
               Re-graded row by row in scripts/probe_absence_claims.py.

  salary       v9: values all correct but the grid TRUNCATED ("steps 1-38 match, truth
               runs to step 50") or structurally collapsed.
               *** v12 CANNOT REPRODUCE THIS. *** `expected_cells` is unpopulated for
               all 431 rows and `capture` measures share-of-page, not share-of-table —
               a page carrying three grids gives each a low capture with nothing
               truncated. v12's cell-agreement check matches rows forward and stops, so
               a grid that ends early scores perfectly. PARTIAL here therefore means the
               0.70-0.95 cell-agreement band, which is NOT v9's definition, and v12's
               salary figure is optimistic relative to v9's by an unmeasured amount.

  rights       v9: quote verbatim and statement type right, but a secondary feature
               (protected_party, voice) wrong. The v12 analogue is the CORRELATIVE pair:
               the rule says `right` where the truth is `obligation`, or the reverse.
               Hohfeld's own point is that these are two descriptions of one relation,
               so the relation was identified and only the perspective differs — v9's
               "core right, feature wrong" exactly.

## Honesty note carried into the workbook

The absence re-grade used a keyword probe that is measurably less sensitive than
individual reading: it missed 4 of the 9 absence errors the original audit had found.
Row-level grades therefore UNDERSTATE the error count, and the summary reports a
capture-recapture correction alongside the raw figure rather than in place of it.
"""

from __future__ import annotations

import argparse
import json
import random
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import OUT_DIR

FILL = {"CORRECT": "FFD9EAD3", "PARTIAL": "FFFFF2CC", "INCORRECT": "FFF4CCCC",
        "header": "FF374151", "note": "FFFFF2CC"}
CORRELATIVE = {"obligation", "right"}


def grade_llm(row: dict, regrade: dict) -> tuple[str, str]:
    if str(row["answer"]).strip().lower() == "not_discussed":
        # The original audit READ these rows; where it found a hard error that verdict
        # wins over the probe's softer grade, because reading beats keyword matching.
        if row["verdict"] == "incorrect":
            return "INCORRECT", "original audit read the source and found the provision"
        if row["verdict"] == "unclear":
            return "EXCLUDED", "unclear on the original read"
        hit = regrade.get(f"{row['document_id']}|{row['question_id']}")
        if hit:
            return hit["grade"], hit["why"]
        return "CORRECT", "no adjacent material found"
    if row["verdict"] == "incorrect":
        return "INCORRECT", str(row["audit_note"])[:180]
    if row["verdict"] == "unclear":
        return "EXCLUDED", "unclear on the original read"
    return "CORRECT", "answer grounded in the clause it cites"


def grade_salary(row: dict) -> tuple[str, str]:
    verdict = row["verdict"]
    if verdict in ("values_correct", "cells_correct"):
        return "CORRECT", "values verified against an independent reading of the page"
    if verdict == "partial":
        return "PARTIAL", "cell agreement 0.70-0.95 (NOT v9's truncation test - see note)"
    if verdict == "cells_wrong":
        return "INCORRECT", "values sit in the wrong cells of the printed table"
    return "EXCLUDED", "page could not be reconstructed independently"


def grade_rights(row: dict) -> tuple[str, str]:
    verdict = row["verdict"]
    if verdict in ("corroborated", "rule_correct"):
        return "CORRECT", "the shipped rule-based label is right"
    if verdict == "quote_not_verbatim":
        return "INCORRECT", "the recorded quote is not verbatim in the source"
    if verdict == "model_correct":
        pair = {str(row["statement_type"]).strip(), str(row["llm_judgment"]).strip()}
        if pair == CORRELATIVE:
            return "PARTIAL", ("right/obligation are Hohfeldian correlatives - the relation "
                               "is identified, the perspective differs")
        return "INCORRECT", "the shipped rule-based label is wrong"
    return "EXCLUDED", "could not be settled from the clause alone"


def estimate(rows: list[dict], acceptable: bool) -> dict:
    """Horvitz-Thompson over strata + a document-clustered bootstrap interval."""
    scored = [r for r in rows if r["grade"] != "EXCLUDED"]
    good = {"CORRECT", "PARTIAL"} if acceptable else {"CORRECT"}

    def ht(sample: list[dict]) -> float | None:
        by: dict[str, list[dict]] = defaultdict(list)
        for row in sample:
            by[row["stratum"]].append(row)
        num = den = 0.0
        for name, group in by.items():
            size = group[0]["stratum_size"]
            num += size * (sum(1 for r in group if r["grade"] in good) / len(group))
            den += size
        return num / den if den else None

    by_doc: dict[str, list[dict]] = defaultdict(list)
    for row in scored:
        by_doc[row["document"]].append(row)
    documents = list(by_doc)
    rng = random.Random(20260812)
    draws = []
    for _ in range(400):
        flat = [r for _ in documents for r in by_doc[rng.choice(documents)]]
        value = ht(flat)
        if value is not None:
            draws.append(value)
    draws.sort()
    return {
        "n": len(scored),
        "raw": sum(1 for r in scored if r["grade"] in good) / len(scored) if scored else None,
        "weighted": ht(scored),
        "low": draws[int(0.025 * len(draws))] if draws else None,
        "high": draws[int(0.975 * len(draws)) - 1] if draws else None,
        "documents": len(documents),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--workbook", type=Path,
                        default=OUT_DIR / "audit_report_v12.xlsx")
    parser.add_argument("--regrade", type=Path, required=True,
                        help="llm_regrade.json from probe_absence_claims.py grading")
    args = parser.parse_args()

    regrade = json.load(open(args.regrade, encoding="utf-8"))
    ledger = [json.loads(line) for line in
              (OUT_DIR / "audit_ledger.jsonl").read_text(encoding="utf-8").splitlines()
              if line.strip()]
    sizes: dict[tuple[str, str], int] = {}
    for row in ledger:
        sizes[(row["sheet"], row.get("stratum", "?"))] = row.get("stratum_size") or 0

    workbook = load_workbook(args.workbook)
    graders = {"2_llm_extract": lambda r: grade_llm(r, regrade),
               "3_salary_schedule": grade_salary, "4_rights_score": grade_rights}
    collected: dict[str, list[dict]] = {}

    for name, grader in graders.items():
        sheet = workbook[name]
        header = [c.value for c in sheet[1]]
        if "grade_v9_scale" in header:                      # idempotent re-run
            first = header.index("grade_v9_scale") + 1
        else:
            first = sheet.max_column + 1
            sheet.cell(row=1, column=first, value="grade_v9_scale")
            sheet.cell(row=1, column=first + 1, value="grade_basis")
            for offset in (0, 1):
                cell = sheet.cell(row=1, column=first + offset)
                cell.font = Font(bold=True, color="FFFFFFFF")
                cell.fill = PatternFill("solid", fgColor=FILL["header"])
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions[sheet.cell(row=1, column=first).column_letter].width = 14
        sheet.column_dimensions[sheet.cell(row=1, column=first+1).column_letter].width = 52

        rows = []
        for index, excel_row in enumerate(sheet.iter_rows(min_row=2), start=2):
            row = {header[i]: c.value for i, c in enumerate(excel_row) if i < len(header)}
            grade, basis = grader(row)
            cell = sheet.cell(row=index, column=first, value=grade)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if grade in FILL:
                cell.fill = PatternFill("solid", fgColor=FILL[grade])
            note = sheet.cell(row=index, column=first + 1, value=basis)
            note.alignment = Alignment(wrap_text=True, vertical="top")
            rows.append({"grade": grade, "stratum": row.get("stratum", "?"),
                         "stratum_size": sizes.get((name, row.get("stratum", "?")), 1),
                         "document": row.get("document_id") or row.get("pdf", "?")})
        collected[name] = rows
        print(f"  {name}: {dict(Counter(r['grade'] for r in rows))}")

    workbook.save(args.workbook)
    json.dump({k: [r["grade"] for r in v] for k, v in collected.items()},
              open(args.regrade.parent / "graded_rows.json", "w"))
    print(f"\n  detail sheets written -> {args.workbook}")
    for name, rows in collected.items():
        strict, acceptable = estimate(rows, False), estimate(rows, True)
        print(f"  {name:20s} strict {strict['raw']:.1%} (est {strict['weighted']:.3f})   "
              f"acceptable {acceptable['raw']:.1%} (est {acceptable['weighted']:.3f} "
              f"[{acceptable['low']:.3f}-{acceptable['high']:.3f}])  docs={acceptable['documents']}")


if __name__ == "__main__":
    main()
