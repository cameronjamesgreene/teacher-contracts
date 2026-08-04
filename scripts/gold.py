"""Import the human audit workbook into a gold-answer table.

`output/audit_report.xlsx` is the only FULL-CENSUS audit in the project: all 106 questions
across 4 documents (Davis, Granite, DCPS, Hawaii DOE), 424 cells, hand-checked against the
source PDFs. Every later audit (v3-v10) samples ~5 questions per document and is badly
biased -- v9 spent 19% of its LLM sample on meta_doc_type_001, the single easiest question
in the bank, and v10 deliberately over-sampled two known-weak questions. Neither can be used
to compare pipeline versions.

TWO PARSING TRAPS, both of which silently destroy a third of the data:

1. The rows are RAGGED -- they carry 7, 8 or 9 populated cells depending on whether the
   evidence and note columns are filled. Index-based column parsing drops rows. This module
   locates the status token by scanning each row instead.
2. `Status` is a JUDGMENT ABOUT A PREDICTION, not a label. For CORRECT rows the gold answer
   is the pipeline's own answer. For PARTIAL/INCORRECT rows the true answer is only in the
   free-text `Audit Note` ("Correct answer is 'yes' with evidence from pages 111-114"), so
   those need extraction rather than a column read.

Anything this module cannot resolve confidently is written with tier='needs_review' rather
than guessed, so a hand pass can finish it without the guesses contaminating the metrics.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import store

ROOT = Path(__file__).resolve().parents[1]
AUDIT_XLSX = ROOT / "output" / "audit_report.xlsx"
STATUSES = ("CORRECT", "PARTIAL", "INCORRECT", "UNVERIFIABLE")

# "Correct answer is 'yes'", "should be yes", "the answer should be 'not_discussed'"
_ANSWER_PAT = [
    re.compile(r"correct answer (?:is|should be)[:\s]+[\"'“]?([\w \-/,\.]+?)[\"'”]?[\.;,]", re.I),
    re.compile(r"should (?:be|have been)[:\s]+[\"'“]?(yes|no|not_discussed|discussed_unclear)[\"'”]?", re.I),
    re.compile(r"\bactual(?:ly)?[:\s]+[\"'“]?(yes|no)[\"'”]?", re.I),
]
_PAGE_PAT = re.compile(r"\bp(?:age|p)?\.?\s*(\d{1,4})", re.I)

# The audited workbook records only a district label, and several of these districts have a
# dozen contracts in the corpus. These are the exact four document_ids of the audited run,
# read from output/1 main_dataset/llm_coding_log.csv -- guessing by district would score the
# wrong contract and silently invalidate every metric built on top.
AUDITED_DOCUMENTS = {
    "davis": "davis_school_district__davis_2019_2020__cc60f6a3",
    "granite": "granite_school_district__professional_agreement_with_gea_2020_2023__7844d12c",
    "dcps": "district_of_columbia_public_schools__dcps_wtu_cba_2023_2028__2af6e7cf",
    "district of columbia": "district_of_columbia_public_schools__dcps_wtu_cba_2023_2028__2af6e7cf",
    "hawaii": "hawaii_department_of_education__2023_2027_hsta_cba_final_07_20_2023__01117dc5",
}
DISTRICT_HINTS = AUDITED_DOCUMENTS


def _rows(sheet):
    for row in sheet.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if any(vals):
            yield vals


def _find_status(vals: list) -> str:
    for v in vals:
        u = v.strip().upper()
        if u in STATUSES:
            return u
    return ""


def _resolve_document_id(district: str) -> str:
    """Map an audit district label to a document_id prefix present in the corpus index."""
    d = district.lower()
    for hint, slug in DISTRICT_HINTS.items():
        if hint in d:
            return slug
    return re.sub(r"[^a-z0-9]+", "_", d).strip("_")


def import_audit(path: Path = AUDIT_XLSX, sheet_name: str = "") -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = sheet_name or next((s for s in wb.sheetnames if "llm" in s.lower()),
                              wb.sheetnames[1])
    sheet = wb[name]
    st = store.get_store()

    counts = {s: 0 for s in STATUSES}
    tiers = {"from_answer": 0, "from_note": 0, "needs_review": 0}
    header_seen = False
    n = 0
    for vals in _rows(sheet):
        status = _find_status(vals)
        if not status:
            continue
        if not header_seen and "question id" in " ".join(vals).lower():
            header_seen = True
            continue
        qid = next((v for v in vals if re.fullmatch(r"[a-z]+_[a-z0-9_]+_\d{3}", v.strip())), "")
        if not qid:
            continue
        district = vals[0] if vals else ""
        document_id = _resolve_document_id(district)
        si = vals.index(next(v for v in vals if v.strip().upper() == status))
        answer = vals[4] if len(vals) > 4 else ""
        note = vals[si + 1] if len(vals) > si + 1 else ""

        if status == "CORRECT":
            gold, tier = answer, "from_answer"
        else:
            gold, tier = "", "needs_review"
            for pat in _ANSWER_PAT:
                m = pat.search(note)
                if m:
                    gold, tier = m.group(1).strip().strip("'\"“”"), "from_note"
                    break
        pm = _PAGE_PAT.search(note or "")
        counts[status] += 1
        tiers[tier] += 1
        with st._lock:
            st._con.execute(
                "INSERT OR REPLACE INTO gold(document_id, qid, gold_answer, gold_page,"
                " gold_quote, source, tier, labeled_at, rationale)"
                " VALUES (?,?,?,?,?,?,?,datetime('now'),?)",
                (document_id, qid, gold, int(pm.group(1)) if pm else None, "",
                 f"{path.name}:{name}:{status}", tier, note[:500]))
        n += 1
    with st._lock:
        st._con.commit()
    return {"rows": n, "status": counts, "tier": tiers}


if __name__ == "__main__":
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else AUDIT_XLSX
    if not src.exists():
        raise SystemExit(f"audit workbook not found: {src}")
    res = import_audit(src)
    print(f"imported {res['rows']} gold cells from {src.name}")
    print(f"  status: {res['status']}")
    print(f"  tier:   {res['tier']}")
    print("\nCells with tier='needs_review' have no machine-recoverable gold answer; the"
          "\ntrue answer is in free-text audit notes and needs one hand pass. They are"
          "\nEXCLUDED from scoring rather than guessed.")
