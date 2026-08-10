#!/usr/bin/env python3
"""Build the single deliverable workbook: four sheets, colour-coded.

One file rather than a scatter of CSVs, because the four things a reader needs are
only meaningful next to each other:

  1. Main dataset      one row per document, 106 questions x 4 fields — the analysis file
  2. Coding log        one row per document-question pair, with quote and page
  3. Citation integrity is every quote real and on the page claimed (no answer key needed)
  4. Accuracy          scores against the documents that have a hand-built answer key

Sheets 3 and 4 measure different things and are deliberately adjacent. Sheet 3 runs on
every document but is **not** an accuracy metric — a pipeline answering "not discussed"
everywhere scores a perfect 1.000 on it, which is why answer coverage sits beside it in
the same table. Sheet 4 is real accuracy but covers only the ~2% of question-document
pairs with verified ground truth.

Colour is used to mean one thing per sheet, never decoratively:
  coding log    answer status — substantive / absent / unclear
  integrity     red where a quote is unverifiable or on the wrong page
  accuracy      green-to-red on the score; grey where the ground truth is stale

Usage:
    python3 scripts/build_workbook.py     # -> output/output_v12/contract_coding_v12.xlsx
"""

from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import subprocess
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import CURRENT_VERSION, WORK, read_codebook

RESULTS = WORK / "output" / "extraction" / "results"
KEYS = WORK / "output" / "extraction" / "keys"
GOLD = WORK / "output" / "extraction" / "answer_gold.csv"
MANIFEST = WORK / "output" / "extraction" / "corpus_manifest.csv"
SCRIPTS = Path(__file__).resolve().parent

# Documents whose answer key was built against extracted text that has since been
# replaced. Their scores are reported but greyed: the key's quotes and pages refer to
# text that no longer exists, so a lower score is not necessarily a worse extraction.
STALE_KEYS = {"pittsburgh"}

ABSENT = {"not_discussed", "discussed_unclear", "not_applicable"}

FILL = {
    "substantive": "FFD9EAD3",   # green
    "absent":      "FFF3F3F3",   # grey
    "unclear":     "FFFCE5CD",   # amber
    "bad":         "FFF4CCCC",   # red
    "ok":          "FFD9EAD3",
    "stale":       "FFE0E0E0",
    "header":      "FF37474F",   # dark slate
}


def clean(value):
    """Strip characters Excel refuses, so one bad document cannot fail the workbook.

    Contract text reaches here from pdftotext and from OCR, and some of it carries
    control characters — a PDF with a broken font encoding produces mojibake
    ("&ULWLFDOLOOQHVV" for "critical illness") mixed with raw control bytes. openpyxl
    raises on those, which would otherwise mean the whole deliverable fails on the
    worst-scanned document in the corpus.
    """
    if not isinstance(value, str):
        return value
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    return ILLEGAL_CHARACTERS_RE.sub("", value)


def status_of(answer: str) -> str:
    value = (answer or "").strip().lower()
    if value == "discussed_unclear":
        return "unclear"
    if value in ABSENT or not value:
        return "absent"
    return "substantive"


def load_manifest() -> dict[str, dict[str, str]]:
    with MANIFEST.open(newline="", encoding="utf-8") as source:
        return {row["document_id"]: row for row in csv.DictReader(source)}


def completed_documents() -> list[str]:
    return sorted(os.path.basename(os.path.dirname(p))
                  for p in glob.glob(str(RESULTS / "*" / "ensemble.jsonl")))


def load_answers(document_id: str) -> dict[str, dict]:
    """Latest answer per question for one document, preferring a subset-polished file."""
    answers: dict[str, dict] = {}
    for name in ("ensemble.jsonl", "final.jsonl"):     # final overrides ensemble
        path = RESULTS / document_id / name
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            for answer in json.loads(line).get("answers", []):
                if answer.get("question_id"):
                    answers[answer["question_id"]] = answer
    return answers


def integrity_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(glob.glob(str(RESULTS / "*" / "citation_audit.csv"))):
        if os.path.basename(os.path.dirname(path)) in ("generalisation", "ocr_check"):
            continue
        with open(path, newline="", encoding="utf-8") as source:
            rows.extend(csv.DictReader(source))
    return rows


def score(document_id: str, gold: Path, label: str) -> dict[str, str] | None:
    """Run the deterministic scorer and parse its one-line summary."""
    ensemble = RESULTS / document_id / "ensemble.jsonl"
    if not ensemble.exists() or not gold.exists():
        return None
    proc = subprocess.run(
        [sys.executable, str(SCRIPTS / "grind_score.py"), "--jsonl", str(ensemble),
         "--label", label, "--gold", str(gold), "--document-id", document_id],
        capture_output=True, text=True)
    line = [l for l in proc.stdout.splitlines() if "overall=" in l]
    if not line:
        return None
    parts = dict(piece.split("=", 1) for piece in line[-1].split() if "=" in piece)
    return {"document": label, **parts}


def accuracy_rows() -> list[dict[str, str]]:
    manifest = load_manifest()

    def find(prefix: str) -> str:
        return next((d for d in manifest if d.startswith(prefix)), "")

    rows = []
    manchester = find("manchester")
    if manchester:
        row = score(manchester, GOLD, "manchester")
        if row:
            row["key"] = "answer_gold.csv (38 q)"
            row["note"] = "DEVELOPMENT SET — the pipeline was tuned against this; biased upward"
            rows.append(row)
    for key in sorted(KEYS.glob("*.csv")):
        document_id = find(key.stem)
        if not document_id:
            continue
        row = score(document_id, key, key.stem)
        if not row:
            continue
        row["key"] = f"keys/{key.name} (16 q)"
        row["note"] = ("STALE KEY — verified against text since replaced by re-OCR"
                       if key.stem in STALE_KEYS else "held out")
        rows.append(row)
    return rows


def build(out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    manifest = load_manifest()
    questions = read_codebook()
    documents = completed_documents()
    workbook = Workbook()

    def style_header(sheet, freeze="A2") -> None:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=FILL["header"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.freeze_panes = freeze
        sheet.row_dimensions[1].height = 30

    # ── 1. main dataset ───────────────────────────────────────────────────────────
    sheet = workbook.active
    sheet.title = "1 main dataset"
    meta = ["document_id", "district", "file_name", "pdf_pages", "text_status"]
    header = meta + [f"{q.qid}_{s}" for q in questions
                     for s in ("answer", "evidence", "page", "confidence")]
    sheet.append(header)
    for document_id in documents:
        row_meta = manifest.get(document_id, {})
        answers = load_answers(document_id)
        row = [document_id, row_meta.get("district", ""), row_meta.get("file_name", ""),
               int(row_meta.get("pdf_pages") or 0), row_meta.get("text_status", "")]
        for question in questions:
            cell = answers.get(question.qid, {})
            row += [cell.get("answer", "not_discussed"), cell.get("evidence", ""),
                    cell.get("page", ""), cell.get("confidence", "")]
        sheet.append([clean(v) for v in row])
    # Colour only the answer columns; evidence/page/confidence inherit meaning from them.
    for column in range(len(meta) + 1, len(header) + 1, 4):
        for row_index in range(2, len(documents) + 2):
            cell = sheet.cell(row=row_index, column=column)
            cell.fill = PatternFill("solid", fgColor=FILL[status_of(str(cell.value))])
    for index, width in enumerate([34, 26, 30, 10, 12], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    style_header(sheet, freeze="F2")

    # ── 2. coding log ─────────────────────────────────────────────────────────────
    sheet = workbook.create_sheet("2 coding log")
    sheet.append(["document_id", "district", "question_id", "topic", "question",
                  "answer", "evidence", "page", "confidence", "status"])
    for document_id in documents:
        district = manifest.get(document_id, {}).get("district", "")
        answers = load_answers(document_id)
        for question in questions:
            cell = answers.get(question.qid, {})
            answer = cell.get("answer", "not_discussed")
            sheet.append([clean(v) for v in [
                document_id, district, question.qid, question.topic,
                question.question, answer, cell.get("evidence", ""),
                cell.get("page", ""), cell.get("confidence", ""), status_of(answer)]])
    for row_index in range(2, sheet.max_row + 1):
        fill = PatternFill("solid", fgColor=FILL[sheet.cell(row=row_index, column=10).value])
        for column in (6, 10):
            sheet.cell(row=row_index, column=column).fill = fill
    for index, width in enumerate([34, 24, 30, 18, 52, 46, 60, 10, 11, 12], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = f"A1:J{sheet.max_row}"
    style_header(sheet, freeze="C2")

    # ── 3. citation integrity ─────────────────────────────────────────────────────
    sheet = workbook.create_sheet("3 citation integrity")
    sheet.append(["district", "pdf_pages", "answered", "questions", "coverage",
                  "verbatim", "contiguous", "on_page", "self_flagged_ungrounded"])
    for row in sorted(integrity_rows(), key=lambda r: r["district"]):
        sheet.append([clean(row["district"]), int(row["pdf_pages"]), int(row["substantive"]),
                      int(row["questions"]), float(row["coverage"]),
                      float(row["verbatim"]), float(row["contiguous"]),
                      float(row["on_page"]), int(row["self_flagged_ungrounded"])])
    for row_index in range(2, sheet.max_row + 1):
        for column in (6, 7, 8):                       # verbatim, contiguous, on_page
            cell = sheet.cell(row=row_index, column=column)
            cell.number_format = "0.000"
            cell.fill = PatternFill(
                "solid", fgColor=FILL["ok"] if float(cell.value) >= 0.99 else FILL["bad"])
        sheet.cell(row=row_index, column=5).number_format = "0.00"
    sheet.append([])
    sheet.append(["NOTE: this is citation integrity, NOT accuracy. A pipeline answering "
                  "'not discussed' everywhere scores 1.000 here — which is why coverage "
                  "is in the same table. Accuracy is sheet 4."])
    sheet.cell(row=sheet.max_row, column=1).font = Font(italic=True)
    for index, width in enumerate([34, 11, 10, 11, 10, 10, 11, 10, 22], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    style_header(sheet)

    # ── 4. accuracy ───────────────────────────────────────────────────────────────
    sheet = workbook.create_sheet("4 accuracy")
    sheet.append(["document", "key", "overall", "status", "required", "completeness",
                  "page", "evidence", "note"])
    for row in accuracy_rows():
        sheet.append([clean(row["document"]), clean(row.get("key", "")), float(row.get("overall", 0)),
                      float(row.get("status", 0)), float(row.get("req", 0)),
                      float(row.get("cov", 0)), float(row.get("page", 0)),
                      float(row.get("ev", 0)), row.get("note", "")])
    for row_index in range(2, sheet.max_row + 1):
        stale = "STALE" in str(sheet.cell(row=row_index, column=9).value)
        dev = "DEVELOPMENT" in str(sheet.cell(row=row_index, column=9).value)
        for column in range(3, 9):
            cell = sheet.cell(row=row_index, column=column)
            cell.number_format = "0.000"
            if stale or dev:
                cell.fill = PatternFill("solid", fgColor=FILL["stale"])
            else:
                cell.fill = PatternFill(
                    "solid",
                    fgColor=FILL["ok"] if float(cell.value) >= 0.80 else FILL["unclear"]
                    if float(cell.value) >= 0.65 else FILL["bad"])
    sheet.append([])
    sheet.append(["NOTE: ground truth covers about 2% of all question-document pairs "
                  "(5 of 42 documents; 38 questions on one, 16 on four). Grey rows are "
                  "not comparable — the development set was tuned against, and a stale "
                  "key was verified against text since replaced. Run-to-run variation is "
                  "about +/-0.05, so a single run cannot resolve smaller differences."])
    sheet.cell(row=sheet.max_row, column=1).font = Font(italic=True)
    for index, width in enumerate([16, 26, 10, 10, 11, 13, 10, 11, 74], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    style_header(sheet)

    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)
    print(f"{len(documents)} documents x {len(questions)} questions -> {out}")
    for name in workbook.sheetnames:
        print(f"  {name}: {workbook[name].max_row - 1} rows")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--out", type=Path,
                        default=WORK / "output" / f"output_{CURRENT_VERSION}"
                                     / f"contract_coding_{CURRENT_VERSION}.xlsx")
    build(parser.parse_args().out)


if __name__ == "__main__":
    main()
