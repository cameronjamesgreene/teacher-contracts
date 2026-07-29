#!/usr/bin/env python3
"""Generate a multi-document verification spreadsheet from llm_coding_log.csv.

Output: output/dcps_verification_2026-06-09.xlsx
Tabs:
  - One tab per document (87 rows each), named by district + years
  - "All Documents" tab with District + File columns prepended (all 174+ rows)
  - "Summary" tab with per-document stats side-by-side

Per-document tabs:
  Question ID | Topic | Question | Answer | Evidence | Page | Confidence |
  Coder Notes | Verified? | Reviewer Notes

All Documents tab prepends:
  District | File | (same fields minus Verified?/Reviewer Notes)
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_DIR   = Path(__file__).resolve().parents[1] / "output"
LOG_CSV   = OUT_DIR / "llm_coding_log.csv"
MAIN_CSV  = OUT_DIR / "llm_main_dataset.csv"
OUT_XLSX  = OUT_DIR / "dcps_verification_2026-06-09.xlsx"

# ── colours ──────────────────────────────────────────────────────────────────
CONF_FILL = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),
    "medium": PatternFill("solid", fgColor="FFEB9C"),
    "low":    PatternFill("solid", fgColor="FFCC99"),
}
HEADER_FILL  = PatternFill("solid", fgColor="2F4F8F")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
TOPIC_FILLS  = [
    PatternFill("solid", fgColor="EBF3FB"),
    PatternFill("solid", fgColor="FFFFFF"),
]
THIN   = Side(style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Column specs for per-document tabs: (header, width, wrap)
DOC_COLUMNS = [
    ("Question ID",      22, False),
    ("Topic",            22, False),
    ("Question",         45, True),
    ("Answer",           35, True),
    ("Evidence",         60, True),
    ("Page",              8, False),
    ("Confidence",       13, False),
    ("Coder Notes",      35, True),
    ("Verified?",        12, False),
    ("Reviewer Notes",   45, True),
]

# Additional leading columns for the All Documents tab
ID_COLUMNS = [
    ("District",         30, False),
    ("File",             42, False),
]
ALL_COLUMNS = ID_COLUMNS + DOC_COLUMNS


def category(qid: str) -> str:
    return qid.split("_")[0]


def _short_tab_name(district: str, start: str, end: str) -> str:
    """Build a ≤31-char Excel tab name from district name + years."""
    # Known abbreviations
    abbrevs = {
        "district of columbia public schools": "DCPS",
        "granite school district": "Granite SD",
    }
    key = district.lower().strip()
    short = abbrevs.get(key, district[:18].strip())
    name = f"{short} {start[-2:]}-{end[-2:]}"
    return name[:31]


def _write_sheet(ws, rows: list[dict], columns: list[tuple],
                 conf_col_idx: int, id_cols: int = 0) -> None:
    """Write header + data rows to worksheet ws."""
    headers = [c[0] for c in columns]
    ws.append(headers)
    for col_idx, (_, width, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    seen_cats: list[str] = []
    def topic_fill(qid: str):
        cat = category(qid)
        if cat not in seen_cats:
            seen_cats.append(cat)
        return TOPIC_FILLS[seen_cats.index(cat) % 2]

    prev_cat = None
    for row_idx, row in enumerate(rows, start=2):
        qid  = row.get("Question ID", "")
        conf = row.get("confidence", "").strip().lower()
        base = topic_fill(qid)

        # Build value list matching columns order
        if id_cols:
            values = [
                row.get("_district", ""),
                row.get("_file", ""),
            ]
        else:
            values = []
        values += [
            qid,
            row.get("topic category", ""),
            row.get("question", ""),
            row.get("answer", ""),
            row.get("evidence", ""),
            row.get("page number", ""),
            row.get("confidence", ""),
            row.get("coder notes", ""),
            "",   # Verified?
            "",   # Reviewer Notes
        ]
        ws.append(values)

        for col_idx, (_, _, wrap) in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=wrap,
                                       shrink_to_fit=False)
            if col_idx == conf_col_idx and conf in CONF_FILL:
                cell.fill = CONF_FILL[conf]
            else:
                cell.fill = base

        ws.row_dimensions[row_idx].height = 60

        # Bold first row of each new category
        if category(qid) != prev_cat:
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)
            prev_cat = category(qid)


def main() -> None:
    # Load district metadata keyed by document_id
    doc_meta: dict[str, dict] = {}
    if MAIN_CSV.exists():
        with MAIN_CSV.open(encoding="utf-8") as f:
            for r in csv.DictReader(f):
                doc_meta[r["document_id"]] = r

    with LOG_CSV.open(encoding="utf-8") as f:
        all_rows = list(csv.DictReader(f))

    # Attach display-name fields used by the All tab
    for row in all_rows:
        meta = doc_meta.get(row.get("document_id", ""), {})
        row["_district"] = meta.get("district_name", row.get("document_id", ""))
        row["_file"]     = meta.get("file_name",     row.get("file_name", ""))

    # Group by document_id preserving insertion order
    from collections import OrderedDict
    by_doc: dict[str, list[dict]] = OrderedDict()
    for row in all_rows:
        by_doc.setdefault(row["document_id"], []).append(row)

    wb = Workbook()
    wb.remove(wb.active)   # remove the default blank sheet

    # ── per-document tabs ────────────────────────────────────────────────────
    conf_col_idx_doc = next(i for i, (h, _, _) in enumerate(DOC_COLUMNS, 1)
                            if h == "Confidence")
    for doc_id, rows in by_doc.items():
        meta = doc_meta.get(doc_id, {})
        tab_name = _short_tab_name(
            meta.get("district_name", doc_id),
            meta.get("start_year", ""),
            meta.get("end_year", ""),
        )
        ws = wb.create_sheet(title=tab_name)
        _write_sheet(ws, rows, DOC_COLUMNS, conf_col_idx=conf_col_idx_doc)

    # ── All Documents tab ────────────────────────────────────────────────────
    conf_col_idx_all = next(i for i, (h, _, _) in enumerate(ALL_COLUMNS, 1)
                            if h == "Confidence")
    ws_all = wb.create_sheet(title="All Documents")
    _write_sheet(ws_all, all_rows, ALL_COLUMNS,
                 conf_col_idx=conf_col_idx_all, id_cols=2)

    # ── Summary tab ──────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    doc_ids = list(by_doc.keys())
    n_docs  = len(doc_ids)

    # Column widths
    ws_sum.column_dimensions["A"].width = 30
    for col_idx in range(2, 2 + n_docs):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = 28

    # Header row: Metric | Doc1 short name | Doc2 short name | ...
    header_row = ["Metric"] + [
        _short_tab_name(
            doc_meta.get(d, {}).get("district_name", d),
            doc_meta.get(d, {}).get("start_year", ""),
            doc_meta.get(d, {}).get("end_year", ""),
        )
        for d in doc_ids
    ]
    ws_sum.append(header_row)
    for cell in ws_sum[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    def _stats(rows: list[dict]) -> dict:
        conf, ans, cat = {}, {}, {}
        for r in rows:
            c = r.get("confidence", "").strip()
            a = r.get("answer", "").strip()
            q = r.get("Question ID", "")
            conf[c] = conf.get(c, 0) + 1
            bucket = a if a in ("not_discussed","discussed_unclear",
                                "not_applicable","ocr_needed") else "substantive"
            ans[bucket] = ans.get(bucket, 0) + 1
            cat[category(q)] = cat.get(category(q), 0) + 1
        return {"conf": conf, "ans": ans, "cat": cat, "n": len(rows)}

    stats_by_doc = {d: _stats(rows) for d, rows in by_doc.items()}

    def _row(label: str, *values) -> list:
        return [label] + list(values)

    def val(d, section, key):
        return str(stats_by_doc[d][section].get(key, 0))

    sections = [
        ("Total questions",         lambda d: str(stats_by_doc[d]["n"])),
        ("",                        lambda d: ""),
        ("── Confidence ──",        lambda d: ""),
        ("  High",                  lambda d: val(d,"conf","high")),
        ("  Medium",                lambda d: val(d,"conf","medium")),
        ("  Low",                   lambda d: val(d,"conf","low")),
        ("",                        lambda d: ""),
        ("── Answer status ──",     lambda d: ""),
        ("  Substantive answer",    lambda d: val(d,"ans","substantive")),
        ("  not_discussed",         lambda d: val(d,"ans","not_discussed")),
        ("  discussed_unclear",     lambda d: val(d,"ans","discussed_unclear")),
        ("  not_applicable",        lambda d: val(d,"ans","not_applicable")),
        ("  ocr_needed",            lambda d: val(d,"ans","ocr_needed")),
        ("",                        lambda d: ""),
        ("── Questions per category ──", lambda d: ""),
    ]
    # Add per-category rows (union of all cats)
    all_cats = sorted({c for d in doc_ids for c in stats_by_doc[d]["cat"]})
    for cat_name in all_cats:
        sections.append((f"  {cat_name}_*", lambda d, cn=cat_name: val(d,"cat",cn)))

    for label, fn in sections:
        ws_sum.append([label] + [fn(d) for d in doc_ids])

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")
    total_rows = sum(len(r) for r in by_doc.values())
    print(f"  {len(by_doc)} document tabs  |  {total_rows} total coding rows  |  {n_docs+2} sheets")


if __name__ == "__main__":
    main()
