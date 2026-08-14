#!/usr/bin/env python3
"""Build audit_report_v12.xlsx — four sheets, in the v9/v11 audit format.

## The two accuracy figures, and why both are needed

The sample is deliberately NOT representative. Absence claims were drawn at 3x their
share, low-fidelity grids at 3x, rule-vs-model disagreements at 3x — because that is
where the errors are, and a uniform sample would have spent its budget confirming the
easy majority. That makes the raw rate on audited rows a *lower bound*, not a corpus
rate, and reporting it alone would understate the pipeline badly.

So `1_Summary` carries both:

  biased (sampled) accuracy   the raw rate over audited rows. Depressed by design.
  estimated corpus accuracy   Horvitz-Thompson: sum(N_h * p_h) / sum(N_h) over strata,
                              using each stratum's true size and the sampling weight
                              recorded when it was drawn. This is the figure that
                              estimates the corpus, and it should be higher.

A bootstrap interval is clustered on DOCUMENTS, not rows: for any claim about contracts
in general the effective sample size is 42, not 4,452.

## What counts as correct, per program

Each program's verdict vocabulary means something different, and flattening them into
one "accuracy" would hide the most important finding.

  2_llm_extract     correct / answer_supported -> right; incorrect -> wrong
  3_salary_schedule values_correct / cells_correct -> right; cells_wrong -> wrong;
                    partial and unverifiable are reported but excluded from the rate
  4_rights_score    the recorded statement_type comes from the DETERMINISTIC rule, so
                    corroborated and rule_correct are right, while model_correct means
                    the stored classification is WRONG. Scoring model_correct as a
                    success would invert the sheet's central finding.

`unclear` is excluded from every rate rather than forced to a side.
"""

from __future__ import annotations

import argparse
import json
import random
import statistics
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import CURRENT_VERSION, OUT_DIR

LEDGER = OUT_DIR / "audit_ledger.jsonl"

# v9's auditor graded CORRECT / PARTIAL / INCORRECT, where PARTIAL meant "right but not
# fully established". Without that middle grade, % Acceptable collapses onto % Strict and
# the two columns say the same thing — which is what happened here first time round.
#
# The tier is recovered from the audit note rather than by re-grading, because the note
# already records WHY each row was accepted. A row accepted on direct evidence (the cited
# quote supports the answer, read) is CORRECT. A row accepted on a hedge — the answer
# synthesises across pages I did not open, or only its internal consistency was checked —
# is PARTIAL: probably right, not independently established. That distinction is the whole
# point of v9's two columns.
# ATTEMPTED AND REVERTED: deriving PARTIAL from the audit note. The notes record how
# STRONGLY each row was verified (read directly / accepted on multi-page synthesis /
# checked only for internal consistency). v9's PARTIAL records something else entirely —
# how COMPLETE the answer is. Mapping one onto the other reclassified 289 of 494 correct
# rows as partial and produced a 41% strict rate and a 0.319 corpus estimate, against 10
# actual errors in 504 scored rows and citation integrity of 0.998. The mapping was
# wrong, not the pipeline. Verification strength is already reported separately as
# "read by Claude" vs "computed"; a real PARTIAL tier needs answers re-graded for
# completeness, which no note in this ledger captures.

RIGHT = {"2_llm_extract": {"correct", "answer_supported"},
         "3_salary_schedule": {"values_correct", "cells_correct"},
         "4_rights_score": {"corroborated", "rule_correct"}}
WRONG = {"2_llm_extract": {"incorrect"},
         "3_salary_schedule": {"cells_wrong"},
         "4_rights_score": {"model_correct", "quote_not_verbatim"}}
EXCLUDED = {"unclear", "unverifiable", "partial"}

FILL = {"right": "FFD9EAD3", "partial": "FFFFF2CC", "wrong": "FFF4CCCC",
        "excluded": "FFF3F3F3",
        "header": "FF374151", "note": "FFFFF2CC"}   # FF374151 = the v9/v11 header slate


def load() -> list[dict]:
    return [json.loads(line) for line in LEDGER.read_text(encoding="utf-8").splitlines()
            if line.strip()]


def outcome(row: dict) -> str:
    verdict = row.get("verdict", "")
    if verdict == "partial":
        return "partial"
    if verdict in RIGHT.get(row["sheet"], set()):
        return "right"
    if verdict in WRONG.get(row["sheet"], set()):
        return "wrong"
    return "excluded"


def rates(rows: list[dict]) -> dict:
    """Biased rate, Horvitz-Thompson estimate, and a document-clustered interval."""
    scored = [r for r in rows if outcome(r) in ("right", "partial", "wrong")]
    if not scored:
        return {"n": 0, "biased": None, "weighted": None, "low": None, "high": None}
    biased = sum(1 for r in scored if outcome(r) == "right") / len(scored)

    # Horvitz-Thompson over strata: each stratum's audited accuracy, weighted by how
    # many rows that stratum actually holds in the corpus.
    by_stratum: dict[str, list[dict]] = defaultdict(list)
    for row in scored:
        by_stratum[row.get("stratum", "?")].append(row)
    total = numerator = 0.0
    for name, group in by_stratum.items():
        size = group[0].get("stratum_size") or len(group)
        accuracy = sum(1 for r in group if outcome(r) == "right") / len(group)
        numerator += size * accuracy
        total += size
    weighted = numerator / total if total else None

    # Bootstrap over DOCUMENTS. Effective n for any statement about contracts is the
    # number of documents, not the number of rows.
    by_document: dict[str, list[dict]] = defaultdict(list)
    for row in scored:
        by_document[row.get("document_id") or row.get("pdf", "?")].append(row)
    documents = list(by_document)
    rng = random.Random(20260812)
    draws = []
    for _ in range(400):
        picked = [by_document[rng.choice(documents)] for _ in documents]
        flat = [r for group in picked for r in group]
        if not flat:
            continue
        n_num = n_den = 0.0
        strata: dict[str, list[dict]] = defaultdict(list)
        for row in flat:
            strata[row.get("stratum", "?")].append(row)
        for name, group in strata.items():
            size = group[0].get("stratum_size") or len(group)
            n_num += size * (sum(1 for r in group if outcome(r) == "right") / len(group))
            n_den += size
        if n_den:
            draws.append(n_num / n_den)
    draws.sort()
    low = draws[int(0.025 * len(draws))] if draws else None
    high = draws[int(0.975 * len(draws)) - 1] if draws else None
    return {"n": len(scored), "biased": biased, "weighted": weighted,
            "low": low, "high": high, "documents": len(documents),
            "by_stratum": {name: (len(g), g[0].get("stratum_size") or len(g),
                                  sum(1 for r in g if outcome(r) == "right") / len(g))
                           for name, g in by_stratum.items()}}


def build(out: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = load()
    workbook = Workbook()

    def style(sheet, widths, freeze=None):
        """Match the v9/v11 workbooks: wrapped, top-aligned cells and rows sized to fit.

        v9 wrapped every data cell and gave each row an explicit height derived from its
        content, so a long audit note is readable in place instead of being clipped to a
        single line. Those workbooks also set no freeze panes.
        """
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            cell.fill = PatternFill("solid", fgColor=FILL["header"])
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.row_dimensions[1].height = 32
        for row_index in range(2, sheet.max_row + 1):
            lines = 1
            for index, width in enumerate(widths, start=1):
                text = sheet.cell(row=row_index, column=index).value
                if isinstance(text, str) and text:
                    lines = max(lines, min(8, -(-len(text) // max(8, int(width) - 1))))
                sheet.cell(row=row_index, column=index).alignment = Alignment(
                    vertical="top", wrap_text=True)
            sheet.row_dimensions[row_index].height = 15 * lines
        if freeze:
            sheet.freeze_panes = freeze

    def clean(value):
        if not isinstance(value, str):
            return value
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
        return ILLEGAL_CHARACTERS_RE.sub("", value)

    # ── 1_Summary ────────────────────────────────────────────────────────────────
    # Laid out like the v9/v11 summaries: a dark banner, scope prose, a colour legend,
    # then one compact table with v9's own columns (Sampled / CORRECT / PARTIAL /
    # INCORRECT / % Acceptable / % Strict) so the two versions can be read side by side.
    # The Horvitz-Thompson estimate is appended as extra columns rather than replacing
    # them, because v9's percentages are raw sample rates and this sample is weighted.
    summary = workbook.active
    summary.title = "1_Summary"

    def banner(text):
        summary.append([text])
        cell = summary.cell(row=summary.max_row, column=1)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=FILL["header"])

    def prose(text, bold=False):
        summary.append([text])
        if bold:
            summary.cell(row=summary.max_row, column=1).font = Font(bold=True)

    counts = {}
    for name in ("2_llm_extract", "3_salary_schedule", "4_rights_score"):
        group = [r for r in rows if r["sheet"] == name]
        tally = Counter(outcome(r) for r in group)
        partial = tally["partial"]
        unver = sum(1 for r in group if r.get("verdict") in ("unverifiable", "unclear"))
        counts[name] = {"rows": len(group), "right": tally["right"], "wrong": tally["wrong"],
                        "partial": partial, "unver": unver, **rates(group)}

    banner(f"v12 Independent Audit — full pipeline (llm_extract + salary_schedule + "
           f"rights_score) on 42 teacher contracts")
    prose("Scope: v12 replaced the v9 question extractor with a retrieval-based pipeline "
          "(fused lexical+dense search, host-derived pages, absence re-verification) and "
          "kept salary_schedule and rights_score essentially unchanged from v9. "
          "salary_schedule gained geometry-first table segmentation; rights_score changed "
          "one line, so appendix pages that are prose are now read instead of skipped.")
    prose("Audit method: 1,470 rows checked. 568 were READ — a source passage was opened "
          "and judged. 902 were COMPUTED — a mechanical check settled them (quote verbatim "
          "on the cited page; grid values matched cell-by-cell against an independent "
          "reconstruction of the page; two independent classifiers agreeing). The two are "
          "counted separately throughout and never pooled.")
    prose("Sampling: deliberately NOT representative. Absence claims, low-fidelity grids "
          "and rule-vs-model disagreements were each drawn at 3x their share, because that "
          "is where errors concentrate. The salary sheet is the entire population (431 "
          "grids), not a sample.")
    # No colour legend. The v9 summary carried one, but here it sat above a table whose
    # own CORRECT/INCORRECT cells are already filled with the same two colours, so it read
    # as a stray band of colour rather than a key - and its swatch mapping was wrong
    # anyway (PARTIAL took the EXCLUDED fill). Removed at the user's direction; the
    # EXCLUDED convention it explained is stated in the note under the table instead.
    summary.append([])

    summary.append(["Program", "Sampled", "CORRECT", "PARTIAL", "INCORRECT",
                    "% Acceptable", "Estimated corpus accuracy",
                    "95% CI low", "95% CI high", "Documents"])
    for cell in summary[summary.max_row]:
        cell.font = Font(bold=True)
    labels = {"2_llm_extract": "LLM extract (questions)",
              "3_salary_schedule": "Salary schedule (grids)",
              "4_rights_score": "Rights (clauses)"}
    for name, label in labels.items():
        c = counts[name]
        scored = c["right"] + c["partial"] + c["wrong"]
        acceptable = (c["right"] + c["partial"]) / scored if scored else 0
        summary.append([label, c["rows"], c["right"], c["partial"], c["wrong"],
                        f"{acceptable:.0%}",
                        round(c["weighted"], 3) if c["weighted"] is not None else "",
                        round(c["low"], 3) if c["low"] is not None else "",
                        round(c["high"], 3) if c["high"] is not None else "",
                        c.get("documents", "")])
        row_index = summary.max_row
        summary.cell(row=row_index, column=3).fill = PatternFill("solid", fgColor=FILL["right"])
        summary.cell(row=row_index, column=5).fill = PatternFill("solid", fgColor=FILL["wrong"])
        summary.cell(row=row_index, column=7).font = Font(bold=True)
    summary.append([])
    prose("% Acceptable is the raw sample rate: rows judged correct, plus PARTIAL where a "
          "partial grade exists. v9 also reported a % Strict column; this audit does not, "
          "because it graded llm_extract and rights_score on a BINARY scale (correct / "
          "incorrect), so Strict would simply repeat Acceptable and invite a false "
          "comparison with v9's stricter figure. Read this column as closer to v9's "
          "% Acceptable than to its % Strict.")
    prose("'Estimated corpus accuracy' is the Horvitz-Thompson figure: each stratum's "
          "audited accuracy weighted by that stratum's true size in the corpus. The "
          "interval is bootstrapped over DOCUMENTS (n=42), not rows, because the effective "
          "sample size for any claim about contracts in general is 42. Rows graded "
          "EXCLUDED - unclear or unverifiable - appear in the detail sheets but are kept "
          "out of every rate above.")
    summary.append([])

    prose("HOW MUCH TO TRUST THE llm_extract FIGURE", bold=True)
    prose("Every one of the 10 errors found in llm_extract came from the 112 rows examined "
          "individually — an 8.9% error rate there and 0% in the 408 rows certified by "
          "cheaper means. That is a property of the method, not the data, so a 40-row "
          "re-check was drawn from the weakest categories and read properly.")
    prose("Result: rows certified only by a keyword probe over an ABSENCE claim were wrong "
          "2 of 14 (14.3%) — the probe's search terms are generated automatically and miss "
          "provisions phrased unexpectedly. Rows certified by answer/quote consistency (0 of "
          "14) and by accepting multi-page synthesis (0 of 12) held up.")
    prose("Correcting for that, the llm_extract estimate falls from 0.986 to about 0.971, "
          "and to 0.959 if the other probe-only category carries the same rate. Treat 0.96 "
          "as the working figure and 0.986 as an upper bound. For comparison, the v9 audit "
          "reported 86% strict — but it was graded by an independent agent against the PDFs, "
          "while this one was graded by the same assistant that built the pipeline, so the "
          "two are not strictly comparable.")
    summary.append([])
    prose("Wrong absences found: 9 confirmed (+2 more in the re-check)   |   "
          "Rights quote-fabrications found: 11   |   Salary grids with cells wrong: 46",
          bold=True)
    summary.append([])

    prose("PER-STRATUM DETAIL — shows how far the sample departs from representative", bold=True)
    summary.append(["Program", "Stratum", "Audited", "Stratum size in corpus",
                    "Accuracy in stratum"])
    for cell in summary[summary.max_row]:
        cell.font = Font(bold=True)
    for name, label in labels.items():
        for stratum, (audited, size, accuracy) in sorted(
                counts[name].get("by_stratum", {}).items()):
            summary.append([label, stratum, audited, size, round(accuracy, 3)])
    summary.append([])

    for line in [
        "FINDINGS",
        "• llm_extract is the strongest program: 0.986 estimated. Its errors are almost "
        "entirely WRONG ABSENCES — a provision that exists, reported as not discussed. "
        "7 of the 9 sit in an appendix line item, a table of contents, or a document index "
        "(Manchester 'Plus.22 Title I Supervisor'; LA 'Prior Experience 163-165'; DeKalb "
        "'EXPERIENCE VERIFICATION FOR SALARY PLACEMENT').",
        "• One case settles the diagnosis: DeKalb's experience-verification section was "
        "cited correctly by one question and marked not_discussed by its sibling. The text "
        "is reachable, so this is per-question retrieval variance, not a retrieval ceiling. "
        "v9's llm_extract had a cross-question reconciliation pass; the v12 pipeline does not.",
        "• rights_score: where the deterministic Table A.3 rule and the model's independent "
        "judgement disagree, the MODEL is right 110 times to 34. The rule fails on negated "
        "modals (may not / cannot / shall not read as permission instead of prohibition), on "
        "passive voice with an implied agent, and on explicit 'shall have the right to'. "
        "Because the stored statement_type comes from the rule, those 110 are recorded "
        "classifications that are wrong, and the bias runs one way — it will skew any "
        "composite rights score until the negation logic is fixed.",
        "• salary_schedule: every grid was checked cell-by-cell against an independent "
        "reconstruction of its page — PDF word geometry for born-digital pages, Apple Vision "
        "at 300 DPI for scans. This is stronger than the v9 audit, which hand-checked a "
        "sample of cells. 103 grids are unverifiable because their page carries several "
        "tables and the reconstruction cannot be attributed to one grid.",
        "",
        "COVERAGE",
        "• rights_score ran on all 42 documents. 41 produced clauses; Anoka-Hennepin "
        "produced none, correctly — it is a 4-page 'Tentative Agreement Summary' of "
        "bargained terms in bullet form rather than clause language, and the same document "
        "answers only 15 of 106 codebook questions.",
        "• The rights sample covers 40 of those 41 documents. Dayton (28 clauses) was not "
        "drawn: with 55,119 clauses and a 519-row sample, a document contributing 28 rows is "
        "unlikely to be selected. A sampling outcome, not a failure.",
        "• Rows in each detail sheet are SHUFFLED across strata. In ledger order they arrive "
        "grouped, which makes the sheet look as though quality collapses partway down. It "
        "does not — every stratum spans all documents.",
        "",
        "WHAT THIS AUDIT DOES NOT ESTABLISH",
        "• Salary lane labels and covered employee groups were never verified — only values.",
        "• A provision on a page nothing cited is invisible to this method.",
        "• Several documents are Memoranda of Agreement or tentative agreements rather than "
        "full contracts (NYC amends specific articles of an underlying CBA). Their low "
        "coverage is correct behaviour and must not be read as extraction failure.",
    ]:
        prose(line, bold=bool(line) and not line.startswith("•"))

    style(summary, [54, 14, 13, 13, 13, 14, 15, 11, 11, 12])

    # ── the three detail sheets ──────────────────────────────────────────────────
    specs = {
        "2_llm_extract": ["document_id", "question_id", "stratum", "page", "answer",
                          "evidence", "verdict", "check_method", "audit_note"],
        "3_salary_schedule": ["pdf", "grid", "pages", "stratum", "cells", "expected_cells",
                              "fidelity", "capture", "page_capture", "verdict",
                              "check_method", "audit_note"],
        "4_rights_score": ["document_id", "stratum", "quote", "statement_type",
                           "llm_judgment", "modal", "negation", "verdict",
                           "check_method", "audit_note"],
    }
    widths = {"2_llm_extract": [30, 28, 15, 12, 58, 58, 14, 14, 70],
              "3_salary_schedule": [26, 40, 9, 13, 8, 10, 9, 9, 11, 14, 13, 68],
              "4_rights_score": [30, 15, 72, 14, 14, 9, 9, 16, 14, 70]}
    for name, fields in specs.items():
        sheet = workbook.create_sheet(name)
        sheet.append(fields)
        # Interleave the strata. The ledger stores rows grouped by stratum, so the sheet
        # opened with 334 consecutive corroborated rows followed by 185 mostly-wrong ones
        # and read as though extraction quality collapsed a third of the way through the
        # corpus. It does not: every stratum spans all documents. Ordering, not data.
        detail = [r for r in rows if r["sheet"] == name]
        random.Random(20260812).shuffle(detail)
        for row in detail:
            sheet.append([clean(row.get(f, "")) for f in fields])
            fill = PatternFill("solid", fgColor=FILL[outcome(row)])
            sheet.cell(row=sheet.max_row, column=fields.index("verdict") + 1).fill = fill
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{sheet.max_row}"
        style(sheet, widths[name])          # v9 sets no freeze panes

    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)
    print(f"  {out}")
    for name in workbook.sheetnames:
        print(f"    {name}: {workbook[name].max_row - 1} rows")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--out", type=Path,
                        default=OUT_DIR / f"audit_report_{CURRENT_VERSION}.xlsx")
    build(parser.parse_args().out)


if __name__ == "__main__":
    main()
