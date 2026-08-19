#!/usr/bin/env python3
"""Probe every audited `not_discussed` answer for material the extractor may have missed.

## Why this exists

The v12 audit graded llm_extract on a BINARY scale (correct / incorrect), while v9 used
three levels. v9's entire strict-vs-acceptable gap — 86% against 95% — came from one
distinction the v12 audit never drew:

    PARTIAL   the document holds material ADJACENT to the question but not squarely on
              point, so `not_discussed` is conservative rather than plainly wrong
              (v9 called these "soft false negatives" — 7 of its 79 rows)
    INCORRECT the document plainly discusses the topic; the absence claim is simply wrong
              (v9's 4 rows)

Every one of v9's 11 non-CORRECT rows was an absence claim. So the two versions can only
be compared by re-grading v12's absence claims on v9's scale. That is what this does.

## The probe

For each absence row, search that document — and only that document — for the question
bank's OWN "Suggested keywords or sections to search" terms, over the FTS5 index built
during extraction. Using the bank's published terms rather than terms invented here keeps
the probe reproducible and independent of whatever the extractor's retrieval did.

A row with no hits is CORRECT with no reading required: nothing in the document matches
the question's own search terms. A row with hits is written out with its passages for a
human (or Claude) to grade, because deciding between ADJACENT and ON POINT is a judgement
about meaning that no keyword count can make. The probe therefore never assigns PARTIAL or
INCORRECT itself — it only decides what must be read.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from utils import OUT_DIR, WORK

INDEX = WORK / "cache" / "contract_search_structural.sqlite3"
BANK = WORK / "extraction_elements_reduced.md"
# Terms this generic carry no signal in a teacher contract - they hit every document and
# would send every row to manual reading. Dropped from probes only, never from grading.
STOPWORDS = {"agreement", "board", "district", "school", "teacher", "teachers", "employee",
             "employees", "contract", "policy", "year", "school year", "staff", "may",
             "shall", "section", "article", "appendix", "page", "state law", "notes"}
MAX_PASSAGES = 4          # per row, ranked by BM25
SNIPPET = 420


def load_bank(path: Path) -> dict[str, dict]:
    """question_id -> {question, keywords, notes} from the markdown table."""
    bank: dict[str, dict] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        if len(cells) < 8 or not re.match(r"^[a-z_]+_\d{3}$", cells[0]):
            continue
        terms = [t.strip().lower() for t in re.split(r"[,;]", cells[5]) if t.strip()]
        bank[cells[0]] = {
            "question": cells[2],
            "keywords": [t for t in terms if t not in STOPWORDS and len(t) > 2],
            "notes": cells[7] if len(cells) > 7 else "",
        }
    return bank


def fts_query(keywords: list[str]) -> str:
    """OR of quoted phrases. Quoting makes multi-word terms phrase matches and keeps
    FTS5 from reading hyphens and periods ('R.S.', 'hard-to-fill') as operators."""
    return " OR ".join('"' + k.replace('"', "") + '"' for k in keywords)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--out", type=Path, required=True)
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    from openpyxl import load_workbook
    sheet = load_workbook(OUT_DIR / "audit_report_v12.xlsx")["2_llm_extract"]
    header = [c.value for c in sheet[1]]
    rows = [{header[i]: c.value for i, c in enumerate(r)} for r in sheet.iter_rows(min_row=2)]
    absence = [r for r in rows if str(r["answer"]).strip().lower() == "not_discussed"]
    if args.limit:
        absence = absence[: args.limit]

    bank = load_bank(BANK)
    connection = sqlite3.connect(str(INDEX))
    out, counts = [], Counter()

    for row in absence:
        qid = str(row["question_id"])
        spec = bank.get(qid)
        if not spec or not spec["keywords"]:
            counts["no_probe_terms"] += 1
            out.append({**{k: row[k] for k in ("document_id", "question_id")},
                        "triage": "no_probe_terms", "question": spec["question"] if spec else "",
                        "hits": []})
            continue
        try:
            found = connection.execute(
                "SELECT p.page_start, p.heading, p.text, bm25(passages_fts) AS score "
                "FROM passages_fts f JOIN passages p ON p.rowid = f.rowid "
                "WHERE passages_fts MATCH ? AND p.document_id = ? "
                "ORDER BY score LIMIT ?",
                (fts_query(spec["keywords"]), row["document_id"], MAX_PASSAGES)).fetchall()
        except sqlite3.OperationalError as error:
            counts["query_error"] += 1
            out.append({"document_id": row["document_id"], "question_id": qid,
                        "triage": "query_error", "question": spec["question"],
                        "hits": [str(error)]})
            continue

        triage = "CORRECT_no_hits" if not found else "needs_reading"
        counts[triage] += 1
        out.append({
            "document_id": row["document_id"], "question_id": qid,
            "question": spec["question"], "notes": spec["notes"][:200],
            "keywords": spec["keywords"], "triage": triage,
            "hits": [{"page": h[0], "heading": (h[1] or "")[:90],
                      "text": re.sub(r"\s+", " ", h[2])[:SNIPPET]} for h in found],
        })

    args.out.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"  absence rows probed: {len(absence)}")
    for key, value in counts.most_common():
        print(f"    {key}: {value}")
    print(f"  -> {args.out}")


if __name__ == "__main__":
    main()
