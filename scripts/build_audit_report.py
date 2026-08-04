#!/usr/bin/env python3
"""Assemble audit_report_v11.xlsx from the three independent audit JSONs.

The v9 and v10 workbooks were produced ad hoc by an AI session with no script behind
them, which is why nothing could reproduce or re-score them and why their sampling
biases went unnoticed. This one is a script: re-run it and you get the same workbook.

Layout matches output_v10/audit_report_v10.xlsx exactly — 1_Summary, 2_llm_extract,
3_salary_schedule, 4_rights_score, with the same column headers — so the two are directly
comparable.

Percentages follow the v10 convention: UNVERIFIABLE rows are excluded from the
denominator, % Acceptable = (CORRECT + PARTIAL) / N, % Strict = CORRECT / N.

    python3 build_audit_report.py --out ../output_v11/audit_report_v11.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

STATUSES = ("CORRECT", "PARTIAL", "INCORRECT", "UNVERIFIABLE")
FILL = {
    "CORRECT": PatternFill("solid", fgColor="C6EFCE"),
    "PARTIAL": PatternFill("solid", fgColor="FFEB9C"),
    "INCORRECT": PatternFill("solid", fgColor="FFC7CE"),
    "UNVERIFIABLE": PatternFill("solid", fgColor="D9D9D9"),
}
HEAD = Font(bold=True)
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")


def tally(items: list) -> dict:
    c = {s: sum(1 for i in items if i.get("status") == s) for s in STATUSES}
    n = len(items) - c["UNVERIFIABLE"]                 # v10 convention
    c["N"] = n
    c["acceptable"] = f"{round(100 * (c['CORRECT'] + c['PARTIAL']) / n)}%" if n else "—"
    c["strict"] = f"{round(100 * c['CORRECT'] / n)}%" if n else "—"
    return c


def _sheet(wb, title: str, headers: list, rows: list, status_col: int, widths: list):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEAD
        cell.fill = HEAD_FILL
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[status_col].fill = FILL.get(str(row[status_col].value), FILL["UNVERIFIABLE"])
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    return ws


def build(llm: dict, salary: dict, rights: dict, out: Path, notes: dict) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tl, ts, tr = tally(llm["items"]), tally(salary["items"]), tally(rights["items"])

    # ── 1_Summary ────────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("1_Summary")
    for line in notes["preamble"]:
        ws.append([line])
    ws.append([])
    ws.append(["Program", "Sampled", "CORRECT", "PARTIAL", "INCORRECT",
               "% Acceptable", "% Strict"])
    for cell in ws[ws.max_row]:
        cell.font = HEAD
        cell.fill = HEAD_FILL
    for label, t in (("LLM extract (questions)", tl),
                     ("Salary schedule (grids)", ts),
                     ("Rights (clauses)", tr)):
        ws.append([label, t["N"], t["CORRECT"], t["PARTIAL"], t["INCORRECT"],
                   t["acceptable"], t["strict"]])
    tot_n = tl["N"] + ts["N"] + tr["N"]
    tot_c = tl["CORRECT"] + ts["CORRECT"] + tr["CORRECT"]
    tot_p = tl["PARTIAL"] + ts["PARTIAL"] + tr["PARTIAL"]
    tot_i = tl["INCORRECT"] + ts["INCORRECT"] + tr["INCORRECT"]
    ws.append(["ALL PROGRAMS", tot_n, tot_c, tot_p, tot_i,
               f"{round(100*(tot_c+tot_p)/tot_n)}%", f"{round(100*tot_c/tot_n)}%"])
    for cell in ws[ws.max_row]:
        cell.font = HEAD
    ws.append([])
    for line in notes["findings"]:
        ws.append([line])
    ws.column_dimensions["A"].width = 118
    for i, w in enumerate([118, 10, 10, 10, 11, 14, 10], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ── 2_llm_extract ────────────────────────────────────────────────────────────────
    _sheet(wb, "2_llm_extract",
           ["District", "Question ID", "Topic", "Question", "Answer", "Cited Page",
            "Status", "Audit Note"],
           [[i.get("district", ""), i.get("qid", ""), i.get("topic", ""),
             i.get("question", ""), i.get("answer", ""), i.get("page", ""),
             i.get("status", ""), i.get("note", "")] for i in llm["items"]],
           status_col=6, widths=[30, 30, 16, 46, 26, 11, 14, 92])

    # ── 3_salary_schedule ────────────────────────────────────────────────────────────
    _sheet(wb, "3_salary_schedule",
           ["District", "Grid File", "Source Page", "Cells Checked", "Cells Wrong",
            "Status", "Audit Note"],
           [[i.get("district", ""), i.get("grid_file", ""), i.get("source_page", ""),
             i.get("cells_checked", 0), i.get("cells_wrong", 0),
             i.get("status", ""), i.get("note", "")] for i in salary["items"]],
           status_col=5, widths=[30, 52, 13, 14, 13, 14, 92])

    # ── 4_rights_score ───────────────────────────────────────────────────────────────
    _sheet(wb, "4_rights_score",
           ["District", "Chunk", "Topic", "Quote", "Statement Type", "LLM Judgment",
            "Acting Party", "Protected Party", "Voice", "Quote Verified",
            "Status", "Audit Note"],
           [[i.get("district", ""), i.get("chunk", ""), i.get("topic", ""),
             i.get("quote", ""), i.get("statement_type", ""), i.get("llm_judgment", ""),
             i.get("acting_party", ""), i.get("protected_party", ""), i.get("voice", ""),
             i.get("quote_verified", ""), i.get("status", ""), i.get("note", "")]
            for i in rights["items"]],
           status_col=10, widths=[26, 8, 16, 60, 15, 14, 13, 15, 10, 13, 14, 88])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--llm", default="/tmp/audit_v11_llm.json")
    ap.add_argument("--salary", default="/tmp/audit_v11_salary.json")
    ap.add_argument("--rights", default="/tmp/audit_v11_rights.json")
    ap.add_argument("--notes", default="", help="JSON with 'preamble' and 'findings' lists")
    ap.add_argument("--out", default=str(Path(__file__).resolve().parents[1]
                                         / "output_v11" / "audit_report_v11.xlsx"))
    args = ap.parse_args()
    llm = json.loads(Path(args.llm).read_text())
    salary = json.loads(Path(args.salary).read_text())
    rights = json.loads(Path(args.rights).read_text())
    notes = json.loads(Path(args.notes).read_text()) if args.notes else \
        {"preamble": ["v11 Independent AI Quality Audit"], "findings": []}
    out = build(llm, salary, rights, Path(args.out), notes)
    print(f"wrote {out}")
    for name, d in (("llm", llm), ("salary", salary), ("rights", rights)):
        t = tally(d["items"])
        print(f"  {name:8s} N={t['N']:3d}  C={t['CORRECT']:3d} P={t['PARTIAL']:3d} "
              f"I={t['INCORRECT']:3d} U={t['UNVERIFIABLE']:3d}  "
              f"acceptable={t['acceptable']:>5s} strict={t['strict']:>5s}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
